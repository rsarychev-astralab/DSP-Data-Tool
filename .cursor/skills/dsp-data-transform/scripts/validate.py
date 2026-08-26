#!/usr/bin/env python3
"""Validate DSP upload xlsx against template field rules."""

import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(2)

HEADERS = [
    "ERID",
    "Номер изначального договора",
    "Дата изначального договора",
    "Тип договора",
    "Предмет договора",
    "Вид деятельности",
    "Тип заказчика",
    "Заказчик",
    "Адрес заказчика",
    "ИНН заказчика или его аналог",
    "Рег.номер заказчика",
    "ОКСМ заказчика",
    "Тип исполнителя",
    "Исполнитель",
    "Адрес исполнителя",
    "ИНН исполнителя или его аналог",
    "Рег.номер исполнителя",
    "ОКСМ исполнителя",
    "Включая НДС",
    "Показы",
    "Сумма",
]
COL_COUNT = len(HEADERS)
# Адреса заказчика/исполнителя (индексы 8 и 14) необязательны.
REQUIRED_ALWAYS = [0, 1, 2, 3, 4, 5, 6, 7, 12, 13, 18, 19, 20]
CUSTOMER_TYPE_IDX = 6
CONTRACTOR_TYPE_IDX = 12
VAT_IDX = 18
IMPRESSIONS_IDX = 19
AMOUNT_IDX = 20

CONTRACT_TYPES = {
    "intermediary", "посреднический",
    "additional", "дополнительноесоглашение", "допсоглашение",
    "original", "оказаниеуслуг", "оказанияуслуг",
}
SUBJECT_TYPES = {
    "representation", "представительство",
    "distribution", "договорнараспространениерекламы",
    "distributionorganization", "договорнаорганизациюраспространениярекламы",
    "mediation", "посредничество",
    "other", "иное", "другое",
}
ACTIVITY_TYPES = {
    "conclude", "заключениедоговоров",
    "distribution", "действиявцеляхраспространениярекламы",
    "commercial", "коммерческоепредставительство",
    "other", "иное", "другое",
    "none", "нет",
}
PARTY_TYPES = {
    "legalperson", "individualentrepreneur", "physicalperson",
    "foreignphysicalperson", "foreignlegalperson",
    "юрлицо", "ип", "физлицо",
    "иностранноефизлицо", "иностранноеюрлицо",
}
FOREIGN_TYPES = {"foreignphysicalperson", "foreignlegalperson", "иностранноефизлицо", "иностранноеюрлицо"}
RU_TYPES = {"legalperson", "individualentrepreneur", "physicalperson", "юрлицо", "ип", "физлицо"}
VAT_VALUES = {"yes", "no"}


def norm(val):
    if val is None:
        return ""
    return str(val).strip()


def norm_key(val):
    return norm(val).lower().replace(" ", "").replace(".", "")


def is_empty(val):
    return norm(val) == ""


def validate_row(row_num, values):
    errors = []
    v = [norm(x) for x in values[:COL_COUNT]]
    while len(v) < COL_COUNT:
        v.append("")

    labels = HEADERS

    for idx in REQUIRED_ALWAYS:
        if is_empty(v[idx]):
            errors.append(f"Row {row_num}: '{labels[idx]}' is required")

    if v[3] and norm_key(v[3]) not in CONTRACT_TYPES:
        errors.append(f"Row {row_num}: invalid 'Тип договора': {v[3]!r}")
    if v[4] and norm_key(v[4]) not in SUBJECT_TYPES:
        errors.append(f"Row {row_num}: invalid 'Предмет договора': {v[4]!r}")
    if v[5] and norm_key(v[5]) not in ACTIVITY_TYPES:
        errors.append(f"Row {row_num}: invalid 'Вид деятельности': {v[5]!r}")

    for idx, name in [(CUSTOMER_TYPE_IDX, "Тип заказчика"), (CONTRACTOR_TYPE_IDX, "Тип исполнителя")]:
        if v[idx] and norm_key(v[idx]) not in PARTY_TYPES:
            errors.append(f"Row {row_num}: invalid '{name}': {v[idx]!r}")

    if v[VAT_IDX] and norm_key(v[VAT_IDX]) not in VAT_VALUES:
        errors.append(f"Row {row_num}: 'Включая НДС' must be yes/no, got {v[VAT_IDX]!r}")

    if v[IMPRESSIONS_IDX]:
        try:
            imps = float(v[IMPRESSIONS_IDX].replace(",", ".").replace(" ", ""))
            if imps < 0 or imps != int(imps):
                errors.append(
                    f"Row {row_num}: 'Показы' must be non-negative integer, got {v[IMPRESSIONS_IDX]!r}"
                )
        except ValueError:
            errors.append(f"Row {row_num}: 'Показы' must be a number, got {v[IMPRESSIONS_IDX]!r}")

    if v[AMOUNT_IDX]:
        try:
            float(v[AMOUNT_IDX].replace(",", ".").replace(" ", ""))
        except ValueError:
            errors.append(f"Row {row_num}: 'Сумма' must be a number, got {v[AMOUNT_IDX]!r}")

    def check_party(party_type_idx, inn_idx, reg_idx, oksm_idx, role):
        ptype = norm_key(v[party_type_idx])
        inn, reg, oksm = v[inn_idx], v[reg_idx], v[oksm_idx]
        if ptype in RU_TYPES and is_empty(inn):
            errors.append(f"Row {row_num}: '{role} ИНН' required for type {v[party_type_idx]!r}")
        if ptype in FOREIGN_TYPES:
            if is_empty(oksm):
                errors.append(f"Row {row_num}: '{role} ОКСМ' required for foreign type")
            if is_empty(inn) and is_empty(reg):
                errors.append(
                    f"Row {row_num}: '{role} ИНН' or '{role} Рег.номер' required for foreign type"
                )

    check_party(CUSTOMER_TYPE_IDX, 9, 10, 11, "Заказчик")
    check_party(CONTRACTOR_TYPE_IDX, 15, 16, 17, "Исполнитель")

    return errors


def main():
    if len(sys.argv) != 2:
        print(f"Usage: {sys.argv[0]} <file.xlsx>", file=sys.stderr)
        sys.exit(2)

    path = Path(sys.argv[1])
    if not path.exists():
        print(f"ERROR: file not found: {path}", file=sys.stderr)
        sys.exit(2)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    all_errors = []
    for col, expected in enumerate(HEADERS, 1):
        actual = norm(ws.cell(1, col).value)
        if actual != expected:
            all_errors.append(f"Column {col}: expected {expected!r}, got {actual!r}")

    data_rows = 0
    for row_num in range(3, ws.max_row + 1):
        values = [ws.cell(row_num, c).value for c in range(1, COL_COUNT + 1)]
        if all(is_empty(x) for x in values):
            continue
        data_rows += 1
        all_errors.extend(validate_row(row_num, values))

    if data_rows == 0:
        all_errors.append("No data rows found (starting from row 3)")

    if all_errors:
        print(f"VALIDATION FAILED ({len(all_errors)} issue(s)):")
        for e in all_errors:
            print(f"  - {e}")
        sys.exit(1)

    print(f"OK: {data_rows} data row(s) validated successfully")
    sys.exit(0)


if __name__ == "__main__":
    main()
