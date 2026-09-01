#!/usr/bin/env python3
"""Transform Seip April source file to DSP upload template."""

import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[4]
TEMPLATE = ROOT / "Шаблон загрузки данных от DSP (2).xlsx"
SOURCE = ROOT / "Исходные данные" / "Сейп апрель.xlsx"
OUTPUT = ROOT / "Результат обработки" / "Сейп апрель.xlsx"

COLUMNS = {
    "erid": 1,
    "contract_no": 2,
    "contract_date": 3,
    "contract_type": 4,
    "contract_subject": 5,
    "activity_type": 6,
    "customer_type": 7,
    "customer_name": 8,
    "customer_address": 9,
    "customer_inn": 10,
    "customer_reg_no": 11,
    "customer_oksm": 12,
    "contractor_type": 13,
    "contractor_name": 14,
    "contractor_address": 15,
    "contractor_inn": 16,
    "contractor_reg_no": 17,
    "contractor_oksm": 18,
    "vat_included": 19,
    "impressions": 20,
    "amount": 21,
}

SOURCE_MAP = {
    "erid": 0,
    "contractor_name": 1,
    "contractor_type": 2,
    "contractor_inn": 3,
    "contractor_reg_no": 4,
    "contractor_oksm": 5,
    "customer_name": 6,
    "customer_type": 7,
    "customer_inn": 8,
    "customer_reg_no": 9,
    "customer_oksm": 10,
    "contract_no": 12,
    "contract_date": 13,
    "contract_type": 14,
    "activity_type": 15,
    "contract_subject": 16,
    "impressions": 17,
    "amount": 18,
    "vat_included": 19,
}

CONTRACT_TYPE = {
    "договороказанияуслуг": "Original",
    "посредническийдоговор": "Intermediary",
    "дополнительноесоглашение": "Additional",
}

SUBJECT_TYPE = {
    "договорнараспространениерекламы": "Distribution",
    "договорнаорганизациюраспространениярекламы": "DistributionOrganization",
    "посредничество": "Mediation",
    "представительство": "Representation",
    "иное": "Other",
}

ACTIVITY_TYPE = {
    "действиявцеляхраспространениярекламы": "Distribution",
    "иное": "Other",
    "заключениедоговоров": "Conclude",
    "коммерческоепредставительство": "Commercial",
}

PARTY_TYPE = {
    "юрлицо": "LegalPerson",
    "ип": "IndividualEntrepreneur",
    "физлицо": "PhysicalPerson",
    "иностранноеюрлицо": "ForeignLegalPerson",
    "иностранноефизлицо": "ForeignPhysicalPerson",
}


def norm_key(val):
    return re.sub(r"[\s\.\-_«»\"']", "", str(val).strip().lower())


def has_value(val):
    if val is None:
        return False
    if isinstance(val, float) and val != val:
        return False
    text = str(val).strip()
    return text != "" and text != "--"


def normalize_text(val):
    if not has_value(val):
        return None
    return str(val).strip()


def normalize_date(val):
    if not has_value(val):
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    text = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def normalize_lookup(val, mapping):
    if not has_value(val):
        return None
    return mapping.get(norm_key(val))


def normalize_vat(val):
    if not has_value(val):
        return None
    key = norm_key(val)
    if key in {"no", "безндс"} or key.startswith("безндс") or "безндс" in key:
        return "no"
    if key in {"yes", "сндс", "ндсда", "22%", "20%", "18%", "10%"}:
        return "yes"
    if "ндс" in key:
        return "yes"
    return None


def normalize_impressions(val):
    if not has_value(val):
        return None
    if isinstance(val, (int, float)):
        num = float(val)
    else:
        num = float(str(val).replace(" ", "").replace(",", "."))
    if num < 0 or num != int(num):
        return None
    return int(num)


def normalize_amount(val):
    if not has_value(val):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    return float(str(val).replace(" ", "").replace(",", "."))


def build_record(row):
    record = {}

    def set_field(key, value):
        if value is not None and has_value(value):
            record[key] = value

    raw = {key: row[idx] if idx < len(row) else None for key, idx in SOURCE_MAP.items()}

    set_field("erid", normalize_text(raw["erid"]))
    set_field("contract_no", normalize_text(raw["contract_no"]))
    set_field("contract_date", normalize_date(raw["contract_date"]))
    set_field("contract_type", normalize_lookup(raw["contract_type"], CONTRACT_TYPE))
    set_field("contract_subject", normalize_lookup(raw["contract_subject"], SUBJECT_TYPE))
    set_field("activity_type", normalize_lookup(raw["activity_type"], ACTIVITY_TYPE))
    set_field("customer_type", normalize_lookup(raw["customer_type"], PARTY_TYPE))
    set_field("customer_name", normalize_text(raw["customer_name"]))
    set_field("customer_inn", normalize_text(raw["customer_inn"]))
    set_field("customer_reg_no", normalize_text(raw["customer_reg_no"]))
    set_field("customer_oksm", normalize_text(raw["customer_oksm"]))
    set_field("contractor_type", normalize_lookup(raw["contractor_type"], PARTY_TYPE))
    set_field("contractor_name", normalize_text(raw["contractor_name"]))
    set_field("contractor_inn", normalize_text(raw["contractor_inn"]))
    set_field("contractor_reg_no", normalize_text(raw["contractor_reg_no"]))
    set_field("contractor_oksm", normalize_text(raw["contractor_oksm"]))
    set_field("vat_included", normalize_vat(raw["vat_included"]))
    set_field("impressions", normalize_impressions(raw["impressions"]))
    set_field("amount", normalize_amount(raw["amount"]))

    return record


def write_cell(ws, row, col, value):
    if has_value(value):
        ws.cell(row, col, value)


def main():
    if not SOURCE.exists():
        print(f"ERROR: source not found: {SOURCE}", file=sys.stderr)
        sys.exit(1)
    if not TEMPLATE.exists():
        print(f"ERROR: template not found: {TEMPLATE}", file=sys.stderr)
        sys.exit(1)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)

    tpl_wb = openpyxl.load_workbook(TEMPLATE)
    tpl_ws = tpl_wb["Sheet1"]
    headers = [tpl_ws.cell(1, c).value for c in range(1, len(COLUMNS) + 1)]
    descriptions = [tpl_ws.cell(2, c).value for c in range(1, len(COLUMNS) + 1)]
    tpl_wb.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for col, value in enumerate(headers, 1):
        if has_value(value):
            ws.cell(1, col, value)
    for col, value in enumerate(descriptions, 1):
        if has_value(value):
            ws.cell(2, col, value)

    src_wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    src_ws = src_wb["Erid"]

    records = []
    skipped_empty = 0
    unknown_contract_type = 0

    for row in src_ws.iter_rows(min_row=2, values_only=True):
        if all(not has_value(v) for v in row):
            skipped_empty += 1
            continue
        record = build_record(row)
        if record:
            records.append(record)
        raw_contract_type = row[SOURCE_MAP["contract_type"]] if len(row) > SOURCE_MAP["contract_type"] else None
        if has_value(raw_contract_type) and normalize_lookup(raw_contract_type, CONTRACT_TYPE) is None:
            unknown_contract_type += 1

    start_row = 3
    for i, record in enumerate(records):
        out_row = start_row + i
        for key, col in COLUMNS.items():
            write_cell(ws, out_row, col, record.get(key))

    wb.save(OUTPUT)
    src_wb.close()

    print(f"Output: {OUTPUT}")
    print(f"Rows written: {len(records)}")
    print(f"Skipped empty rows: {skipped_empty}")
    print(f"Rows with unknown contract type (field skipped): {unknown_contract_type}")


if __name__ == "__main__":
    main()
