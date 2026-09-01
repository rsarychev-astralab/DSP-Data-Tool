"""Метчинг строк шаблона DSP с «База договоров»."""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl

from app.config import (
    CONTRACT_ATTRS_CANDIDATES,
    CONTRACT_ATTRS_SHEET,
    resolve_contract_attrs_path as _resolve_contract_attrs_path,
)
from app.engine.lookups import TEMPLATE_COL_COUNT, TEMPLATE_COLUMNS
from app.engine.normalize import has_value
from app.engine.template import TEXT_NUMBER_FORMAT, write_cell
from app.matching.keys import build_match_key
from app.validation.validate import COL_FIELD
from app.validation.validate import HEADERS as TEMPLATE_HEADERS
from app.validation.validate import is_empty

DATA_START_ROW = 3
OUTPUT_HEADERS = ["ERID", "ID в OTM", "ID в OZON", "ID в VK"]
OUTPUT_SHEET_IDS = "ID"
OUTPUT_SHEET_UPLOAD = "Выгрузка с ID"
MATCH_ID_HEADERS = ("ID в OTM", "ID в OZON", "ID в VK")

# Колонки листа «Выгрузка с ID»: (заголовок, ключ поля в строке шаблона)
UPLOAD_DETAIL_COLUMNS: tuple[tuple[str, str], ...] = (
    ("ERID", "erid"),
    ("Номер изначального договора", "contract_no"),
    ("Дата изначального договора", "contract_date"),
    ("Тип договора", "contract_type"),
    ("Предмет договора", "contract_subject"),
    ("Вид деятельности", "activity_type"),
    ("Тип заказчика", "customer_type"),
    ("Заказчик", "customer_name"),
    ("Адрес заказчика", "customer_address"),
    ("ИНН заказчика или его аналог", "customer_inn"),
    ("Рег.номер заказчика", "customer_reg_no"),
    ("ОКСМ заказчика", "customer_oksm"),
    ("Тип исполнителя", "contractor_type"),
    ("Исполнитель", "contractor_name"),
    ("Адрес исполнителя", "contractor_address"),
    ("ИНН исполнителя или его аналог", "contractor_inn"),
    ("Рег.номер исполнителя", "contractor_reg_no"),
    ("ОКСМ исполнителя", "contractor_oksm"),
)

_UPLOAD_FIELD_INDEX = {field: idx for idx, field in enumerate(COL_FIELD)}

# Заголовки листа OriginalContract в «База договоров.xlsx»
DB_HEADERS = {
    "contract_no": "Номер",
    "contract_date": "Дата",
    "contractor_name": "Исполнитель",
    "contractor_inn": "ИНН исполнителя",
    "customer_name": "Заказчик",
    "customer_inn": "ИНН заказчика",
    "contract_type": "Тип договора",
    "contract_subject": "Сведения о предмете договора",
    "activity_type": "Вид деятельности",
    "inactive": "Неактуален",
    "id_otm": "ID в OTM",
    "id_ozon": "ID в OZON",
    "id_vk": "ID в VK",
}

# Колонки выгрузки шаблона для метчинга (0-based)
UPLOAD_COL = {
    key: TEMPLATE_COLUMNS[key] - 1
    for key in (
        "erid",
        "contract_no",
        "contract_date",
        "contract_type",
        "contract_subject",
        "customer_inn",
        "contractor_inn",
    )
}


@dataclass(frozen=True)
class ContractEntry:
    id_otm: Any
    id_ozon: Any
    id_vk: Any


@dataclass(frozen=True)
class MatchResult:
    output_bytes: bytes
    output_filename: str
    rows_total: int
    rows_matched: int
    rows_unmatched: int
    contracts_total: int
    contracts_matched: int
    contracts_unmatched: int


def resolve_contract_attrs_path() -> Path | None:
    return _resolve_contract_attrs_path()


def build_matched_filename(original_filename: str | None) -> str:
    if original_filename:
        stem = Path(original_filename).stem
        if stem.endswith("_matched"):
            return f"{stem}.xlsx"
        return f"{stem}_matched.xlsx"
    return "dsp_matched.xlsx"


def _cell(row: tuple, index: int) -> Any:
    return row[index] if index < len(row) else None


def _is_inactive(val) -> bool:
    if not has_value(val):
        return False
    return norm_key_inactive(str(val).strip().lower())


def norm_key_inactive(text: str) -> bool:
    return text in {"да", "yes", "true", "1", "y"}


def _header_columns(ws, titles: dict[str, str]) -> dict[str, int]:
    by_title: dict[str, int] = {}
    for col in range(1, (ws.max_column or 0) + 1):
        val = ws.cell(1, col).value
        if has_value(val):
            by_title[str(val).strip()] = col
    columns: dict[str, int] = {}
    missing = []
    for field, title in titles.items():
        if title in by_title:
            columns[field] = by_title[title]
        elif field not in ("id_otm", "id_ozon", "id_vk"):
            missing.append(title)
    for id_field, title in (
        ("id_otm", "ID в OTM"),
        ("id_ozon", "ID в OZON"),
        ("id_vk", "ID в VK"),
    ):
        if title in by_title:
            columns[id_field] = by_title[title]
    if missing:
        raise ValueError(
            "В справочнике договоров нет колонок: " + ", ".join(f"«{t}»" for t in missing)
        )
    return columns


def _format_id(val) -> str | None:
    if not has_value(val):
        return None
    if isinstance(val, float) and val == val and val == int(val):
        return str(int(val))
    if isinstance(val, int):
        return str(val)
    return str(val).strip()


def _load_contract_index(path: Path) -> dict[tuple[str, ...], ContractEntry]:
    """Load contract DB into match key index (first row wins on duplicate keys)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = (
            CONTRACT_ATTRS_SHEET
            if CONTRACT_ATTRS_SHEET in wb.sheetnames
            else wb.sheetnames[0]
        )
        ws = wb[sheet_name]
        cols = _header_columns(ws, DB_HEADERS)
        index: dict[tuple[str, ...], ContractEntry] = {}

        def get_col(row_vals: list, field: str):
            col = cols[field]
            return row_vals[col - 1] if col - 1 < len(row_vals) else None

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            row_vals = list(row)
            if not has_value(get_col(row_vals, "contract_no")):
                continue
            if "inactive" in cols and _is_inactive(get_col(row_vals, "inactive")):
                continue

            key = build_match_key(
                contract_no=get_col(row_vals, "contract_no"),
                contract_date=get_col(row_vals, "contract_date"),
                customer_inn=get_col(row_vals, "customer_inn"),
                contractor_inn=get_col(row_vals, "contractor_inn"),
                contract_type=get_col(row_vals, "contract_type"),
                contract_subject=get_col(row_vals, "contract_subject"),
                from_db=True,
            )
            if key in index:
                continue
            index[key] = ContractEntry(
                id_otm=_format_id(get_col(row_vals, "id_otm")) if "id_otm" in cols else None,
                id_ozon=_format_id(get_col(row_vals, "id_ozon")) if "id_ozon" in cols else None,
                id_vk=_format_id(get_col(row_vals, "id_vk")) if "id_vk" in cols else None,
            )
        return index
    finally:
        wb.close()


def _validate_transform_output(ws) -> None:
    """Метчинг принимает только файл после преобразования, не сырую выгрузку партнёра."""
    for col, expected in enumerate(TEMPLATE_HEADERS, 1):
        actual = ws.cell(1, col).value
        if not has_value(actual):
            raise ValueError(
                "Загрузите файл после преобразования с вкладки «Загрузка и преобразование». "
                f"В строке заголовков не хватает колонки «{expected}»."
            )
        if str(actual).strip() != expected:
            raise ValueError(
                "Загрузите файл после преобразования с вкладки «Загрузка и преобразование», "
                "а не исходную выгрузку DSP. "
                f"Ожидался заголовок «{expected}» (колонка {col}), получено «{actual}»."
            )


def _read_upload_rows(data: bytes) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb.active
        _validate_transform_output(ws)
        rows: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=DATA_START_ROW, max_col=TEMPLATE_COL_COUNT, values_only=True):
            values = list(row) if row else []
            if all(is_empty(x) for x in values):
                continue
            rows.append({
                "cells": values,
                "erid": _cell(values, UPLOAD_COL["erid"]),
                "contract_no": _cell(values, UPLOAD_COL["contract_no"]),
                "contract_date": _cell(values, UPLOAD_COL["contract_date"]),
                "contract_type": _cell(values, UPLOAD_COL["contract_type"]),
                "contract_subject": _cell(values, UPLOAD_COL["contract_subject"]),
                "customer_inn": _cell(values, UPLOAD_COL["customer_inn"]),
                "contractor_inn": _cell(values, UPLOAD_COL["contractor_inn"]),
            })
        if not rows:
            raise ValueError("В файле нет строк данных (ожидаются с 3-й строки)")
        return rows
    finally:
        wb.close()


def _write_ids_sheet(ws, summary_rows: list[tuple]) -> None:
    for col, title in enumerate(OUTPUT_HEADERS, 1):
        cell = ws.cell(1, col, title)
        cell.number_format = TEXT_NUMBER_FORMAT
    for i, (erid, id_otm, id_ozon, id_vk) in enumerate(summary_rows, start=DATA_START_ROW):
        for col, val in enumerate((erid, id_otm, id_ozon, id_vk), 1):
            if val is None:
                continue
            cell = ws.cell(i, col, str(val))
            cell.number_format = TEXT_NUMBER_FORMAT


def _write_upload_sheet(ws, detailed_rows: list[tuple]) -> None:
    """Выбранные поля выгрузки + ID в OTM/OZON/VK."""
    for col, (title, _) in enumerate(UPLOAD_DETAIL_COLUMNS, 1):
        cell = ws.cell(1, col, title)
        cell.number_format = TEXT_NUMBER_FORMAT
    id_col_start = len(UPLOAD_DETAIL_COLUMNS) + 1
    for col_offset, title in enumerate(MATCH_ID_HEADERS):
        cell = ws.cell(1, id_col_start + col_offset, title)
        cell.number_format = TEXT_NUMBER_FORMAT

    for i, (cells, id_otm, id_ozon, id_vk) in enumerate(detailed_rows, start=DATA_START_ROW):
        row_cells = list(cells) if cells else []
        for out_col, (_, field_key) in enumerate(UPLOAD_DETAIL_COLUMNS, 1):
            src_idx = _UPLOAD_FIELD_INDEX[field_key]
            val = row_cells[src_idx] if src_idx < len(row_cells) else None
            write_cell(ws, i, out_col, field_key, val)
        for col_offset, val in enumerate((id_otm, id_ozon, id_vk)):
            if val is None:
                continue
            cell = ws.cell(i, id_col_start + col_offset, str(val))
            cell.number_format = TEXT_NUMBER_FORMAT


def _write_output(
    summary_rows: list[tuple],
    detailed_rows: list[tuple],
) -> bytes:
    wb = openpyxl.Workbook()
    ws_ids = wb.active
    ws_ids.title = OUTPUT_SHEET_IDS
    _write_ids_sheet(ws_ids, summary_rows)

    ws_upload = wb.create_sheet(OUTPUT_SHEET_UPLOAD)
    _write_upload_sheet(ws_upload, detailed_rows)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


_CONTRACT_FIELDS = (
    "contract_no",
    "contract_date",
    "contract_type",
    "contract_subject",
    "customer_inn",
    "contractor_inn",
)


def _contract_group_id(row: dict[str, Any]) -> str:
    """Один договор = полный ключ (номер, дата, оба ИНН, тип, предмет), не только номер."""
    key = build_match_key(
        contract_no=row["contract_no"],
        contract_date=row["contract_date"],
        customer_inn=row["customer_inn"],
        contractor_inn=row["contractor_inn"],
        contract_type=row["contract_type"],
        contract_subject=row["contract_subject"],
        from_db=False,
    )
    return "key:" + "|".join(key)


def _canonical_contract_fields(rows: list[dict[str, Any]]) -> dict[str, Any]:
    """Собирает атрибуты договора из всех строк группы (первое непустое значение)."""
    canonical: dict[str, Any] = {}
    for field in _CONTRACT_FIELDS:
        for row in rows:
            val = row.get(field)
            if has_value(val):
                canonical[field] = val
                break
    return canonical


def _match_contract_groups(
    upload_rows: list[dict[str, Any]],
    index: dict[tuple[str, ...], ContractEntry],
) -> dict[str, ContractEntry | None]:
    """Сметчить уникальные договоры в файле с базой, не каждую строку ERID отдельно."""
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in upload_rows:
        groups[_contract_group_id(row)].append(row)

    resolved: dict[str, ContractEntry | None] = {}
    for group_id, group_rows in groups.items():
        canonical = _canonical_contract_fields(group_rows)
        key = build_match_key(
            contract_no=canonical.get("contract_no"),
            contract_date=canonical.get("contract_date"),
            customer_inn=canonical.get("customer_inn"),
            contractor_inn=canonical.get("contractor_inn"),
            contract_type=canonical.get("contract_type"),
            contract_subject=canonical.get("contract_subject"),
            from_db=False,
        )
        resolved[group_id] = index.get(key)
    return resolved


def _apply_matching(
    upload_bytes: bytes, attrs_path: Path
) -> tuple[bytes, int, int, int, int, int, int]:
    index = _load_contract_index(attrs_path)
    upload_rows = _read_upload_rows(upload_bytes)
    group_entries = _match_contract_groups(upload_rows, index)

    groups_seen: set[str] = set()
    contracts_matched = 0
    summary_rows: list[tuple] = []
    detailed_rows: list[tuple] = []
    rows_matched = 0

    for row in upload_rows:
        group_id = _contract_group_id(row)
        entry = group_entries[group_id]
        erid = _format_id(row["erid"]) or str(row["erid"]).strip()
        if entry:
            rows_matched += 1
            ids = (entry.id_otm, entry.id_ozon, entry.id_vk)
        else:
            ids = (None, None, None)
        summary_rows.append((erid, *ids))
        detailed_rows.append((row.get("cells") or [], *ids))
        if group_id not in groups_seen:
            groups_seen.add(group_id)
            if entry:
                contracts_matched += 1

    rows_total = len(upload_rows)
    contracts_total = len(group_entries)
    contracts_unmatched = contracts_total - contracts_matched
    return (
        _write_output(summary_rows, detailed_rows),
        rows_total,
        rows_matched,
        rows_total - rows_matched,
        contracts_total,
        contracts_matched,
        contracts_unmatched,
    )


def match_workbook_bytes(
    upload_bytes: bytes,
    *,
    original_filename: str | None = None,
) -> MatchResult:
    attrs_path = resolve_contract_attrs_path()
    if attrs_path is None:
        expected = ", ".join(f"«{p.name}»" for p in CONTRACT_ATTRS_CANDIDATES)
        raise FileNotFoundError(
            "Не найден файл атрибутов договоров в папке «Справка». "
            f"Добавьте один из файлов: {expected}"
        )

    (
        output_bytes,
        rows_total,
        rows_matched,
        rows_unmatched,
        contracts_total,
        contracts_matched,
        contracts_unmatched,
    ) = _apply_matching(upload_bytes, attrs_path)
    return MatchResult(
        output_bytes=output_bytes,
        output_filename=build_matched_filename(original_filename),
        rows_total=rows_total,
        rows_matched=rows_matched,
        rows_unmatched=rows_unmatched,
        contracts_total=contracts_total,
        contracts_matched=contracts_matched,
        contracts_unmatched=contracts_unmatched,
    )
