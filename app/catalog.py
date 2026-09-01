import re
from dataclasses import dataclass
from functools import lru_cache

import openpyxl

from app.config import SPRAVKA_DSP_PATH
from app.profiles.loader import has_transform_profile, resolve_profile_id

_ID_RE = re.compile(r"[^a-z0-9_]+")


def _slug(value: str) -> str:
    text = str(value).strip().lower().split(",")[0].strip()
    slug = _ID_RE.sub("_", text).strip("_")
    return slug or "unknown"


@dataclass(frozen=True)
class DspCatalogEntry:
    id: str
    display_name: str
    dsp_ids: list[str]
    contract: str | None
    report_to_ord: bool | None
    url: str | None
    has_profile: bool
    profile_id: str | None


@dataclass(frozen=True)
class CatalogSnapshot:
    entries: tuple[DspCatalogEntry, ...]
    warnings: tuple[str, ...]


def _parse_report(val) -> bool | None:
    if val is None or str(val).strip() == "":
        return None
    key = str(val).strip().lower()
    if key in {"да", "yes", "true", "1"}:
        return True
    if key in {"нет", "no", "false", "0"}:
        return False
    return None


def _resolve_catalog_id(dsp_ids: list[str], display_name: str, seen: set[str]) -> str:
    """Уникальный id партнёра: dsp id, иначе slug названия, при коллизии — суффикс."""
    candidates: list[str] = []
    if dsp_ids:
        candidates.append(_slug(dsp_ids[0]))
    if display_name:
        name_slug = _slug(display_name)
        if name_slug not in candidates:
            candidates.append(name_slug)
    if not candidates:
        candidates.append("unknown")

    for candidate in candidates:
        if candidate not in seen:
            return candidate

    base = candidates[0]
    n = 2
    while f"{base}_{n}" in seen:
        n += 1
    return f"{base}_{n}"


def _build_catalog_snapshot() -> CatalogSnapshot:
    path = SPRAVKA_DSP_PATH
    if not path.exists():
        return CatalogSnapshot((), ())

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["DSP"] if "DSP" in wb.sheetnames else wb.active

    entries: list[DspCatalogEntry] = []
    seen: set[str] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(cell is not None and str(cell).strip() for cell in row):
            continue
        display_name = str(row[0]).strip() if row[0] is not None else ""
        dsp_id_raw = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if not display_name and not dsp_id_raw:
            continue
        dsp_ids = [p.strip() for p in dsp_id_raw.split(",") if p.strip()]
        catalog_id = _resolve_catalog_id(dsp_ids, display_name, seen)
        seen.add(catalog_id)

        profile_id = resolve_profile_id(catalog_id)
        has_profile = has_transform_profile(catalog_id)
        contract = str(row[2]).strip() if len(row) > 2 and row[2] is not None else None
        url = str(row[4]).strip() if len(row) > 4 and row[4] is not None else None
        if contract == "":
            contract = None
        if url == "":
            url = None
        entries.append(
            DspCatalogEntry(
                id=catalog_id,
                display_name=display_name or catalog_id,
                dsp_ids=dsp_ids or [catalog_id],
                contract=contract,
                report_to_ord=_parse_report(row[3] if len(row) > 3 else None),
                url=url,
                has_profile=has_profile,
                profile_id=profile_id if has_profile else None,
            )
        )
    wb.close()
    return CatalogSnapshot(tuple(entries), ())


@lru_cache(maxsize=1)
def _catalog_snapshot() -> CatalogSnapshot:
    return _build_catalog_snapshot()


def reload_dsp_catalog() -> None:
    _catalog_snapshot.cache_clear()


def load_dsp_catalog() -> list[DspCatalogEntry]:
    return list(_catalog_snapshot().entries)


def get_catalog_warnings() -> list[str]:
    return list(_catalog_snapshot().warnings)


def get_catalog_entry(partner_id: str) -> DspCatalogEntry | None:
    resolved = resolve_profile_id(partner_id)
    for entry in _catalog_snapshot().entries:
        if entry.id == partner_id or entry.id == resolved:
            return entry
        if entry.profile_id and entry.profile_id == resolved:
            return entry
    return None
