from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from pathlib import Path

import xlrd
import xlwt
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook

from app.fl_address.core import resolve_fl_address

MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_BATCH_ROWS = 50_000
ALLOWED_INPUT = {".xlsx", ".xlsm", ".xls", ".csv", ".txt", ".tsv"}
ALLOWED_OUTPUT = {"csv", "xls", "xlsx"}
JOB_TTL_SEC = 60 * 60
INN_HEADERS = {"инн", "inn"}
ADDRESS_HEADERS = {"адрес", "адрес физического лица", "предлагаемый адрес"}
ADDRESS_HEADER = "Адрес"
RESULT_NAME = "fl_address_result"
MAX_PROBLEMS_IN_STATUS = 100
MEDIA_TYPES = {
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "xls": "application/vnd.ms-excel",
    "csv": "text/csv; charset=utf-8",
}


@dataclass
class BatchProblem:
    row: int
    inn: str
    error: str


@dataclass
class BatchStats:
    total: int = 0
    filled: int = 0
    errors: int = 0
    empty: int = 0
    problems: list[BatchProblem] = field(default_factory=list)


@dataclass
class BatchFileResult:
    content: bytes
    media_type: str
    filename: str
    output_format: str
    stats: BatchStats


def suffix_allowed(filename: str) -> bool:
    return Path(filename).suffix.lower() in ALLOWED_INPUT


def default_output_format(filename: str) -> str:
    lower = (filename or "").lower()
    if lower.endswith((".xlsx", ".xlsm")):
        return "xlsx"
    if lower.endswith(".xls"):
        return "xls"
    return "csv"


def resolve_output_format(filename: str, output_format: str) -> str:
    fmt = (output_format or "auto").strip().lower()
    if fmt == "auto":
        fmt = default_output_format(filename)
    if fmt not in ALLOWED_OUTPUT:
        raise HTTPException(
            status_code=400,
            detail="Формат результата: xlsx, xls или csv",
        )
    return fmt


def problems_payload(problems: list[BatchProblem]) -> dict:
    items = [
        {"row": p.row, "inn": p.inn, "error": p.error}
        for p in problems[:MAX_PROBLEMS_IN_STATUS]
    ]
    return {"items": items, "total": len(problems)}


def _header_key(value: object | None) -> str:
    return str(value or "").strip().lower().replace("ё", "е")


def _cell_empty(value: object | None) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and value != value:
        return True
    return str(value).strip() == ""


def _row_blank(row: list[object]) -> bool:
    return all(_cell_empty(cell) for cell in row)


def find_inn_column(rows: list[list[object]]) -> tuple[int, int]:
    for row_idx, row in enumerate(rows[:10]):
        for col_idx, cell in enumerate(row):
            if _header_key(cell) in INN_HEADERS:
                return row_idx, col_idx
    raise HTTPException(status_code=400, detail="В файле нет колонки «ИНН»")


def enrich_rows(rows: list[list[object]]) -> tuple[list[list[object]], BatchStats]:
    if not rows:
        raise HTTPException(status_code=400, detail="Файл пустой")

    header_row_idx, inn_col = find_inn_column(rows)
    header = rows[header_row_idx]
    next_idx = inn_col + 1
    reuse = next_idx < len(header) and _header_key(header[next_idx]) in ADDRESS_HEADERS
    if not reuse:
        for row_idx, row in enumerate(rows):
            while len(row) <= inn_col:
                row.append(None)
            row.insert(inn_col + 1, ADDRESS_HEADER if row_idx == header_row_idx else None)
        addr_col = inn_col + 1
    else:
        addr_col = next_idx
        if _cell_empty(header[addr_col]):
            header[addr_col] = ADDRESS_HEADER

    data_rows = rows[header_row_idx + 1 :]
    if len(data_rows) > MAX_BATCH_ROWS:
        raise HTTPException(
            status_code=400,
            detail=f"Слишком много строк: {len(data_rows)}. Лимит: {MAX_BATCH_ROWS}",
        )

    stats = BatchStats()
    for offset, row in enumerate(data_rows):
        sheet_row = header_row_idx + 2 + offset
        if _row_blank(row):
            continue
        while len(row) <= max(inn_col, addr_col):
            row.append(None)
        inn_val = row[inn_col]
        result = resolve_fl_address(inn_val)
        stats.total += 1
        if not result["inn"] and not result["error"]:
            stats.empty += 1
            continue
        if result["error"]:
            row[addr_col] = result["error"]
            stats.errors += 1
            stats.problems.append(
                BatchProblem(row=sheet_row, inn=result["inn"], error=result["error"])
            )
        else:
            row[addr_col] = result["address"]
            stats.filled += 1

    if stats.total == 0:
        raise HTTPException(status_code=400, detail="В файле нет строк с данными после заголовка")

    return rows, stats


def _decode_text(content: bytes) -> str:
    try:
        return content.decode("utf-8-sig")
    except UnicodeDecodeError:
        return content.decode("cp1251", errors="replace")


def _detect_delimiter(text: str) -> str:
    sample = text[:4096]
    tab_count = sample.count("\t")
    semicolon_count = sample.count(";")
    comma_count = sample.count(",")
    if tab_count > semicolon_count and tab_count > comma_count:
        return "\t"
    if comma_count > semicolon_count:
        return ","
    return ";"


def parse_text_table(content: bytes) -> tuple[list[list[object]], str]:
    text = _decode_text(content).strip()
    if not text:
        raise HTTPException(status_code=400, detail="Файл пустой")
    delimiter = _detect_delimiter(text)
    rows = [list(row) for row in csv.reader(io.StringIO(text), delimiter=delimiter)]
    return rows, delimiter


def parse_xlsx_table(content: bytes) -> list[list[object]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Не удалось прочитать XLSX: {exc}") from exc
    try:
        sheet = workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def parse_xls_table(content: bytes) -> list[list[object]]:
    try:
        book = xlrd.open_workbook(file_contents=content)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Не удалось прочитать XLS: {exc}") from exc
    sheet = book.sheet_by_index(0)
    return [sheet.row_values(row_idx) for row_idx in range(sheet.nrows)]


def rows_to_csv(rows: list[list[object]], delimiter: str) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=delimiter, lineterminator="\n")
    for row in rows:
        writer.writerow(["" if cell is None else cell for cell in row])
    return ("\ufeff" + buffer.getvalue()).encode("utf-8")


def rows_to_xlsx(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Результат"
    for row in rows:
        sheet.append(["" if cell is None else cell for cell in row])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def rows_to_xls(rows: list[list[object]]) -> bytes:
    workbook = xlwt.Workbook(encoding="utf-8")
    sheet = workbook.add_sheet("Результат")
    for row_idx, row in enumerate(rows):
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            sheet.write(row_idx, col_idx, cell)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def output_filename(output_format: str) -> str:
    return f"{RESULT_NAME}.{output_format}"


def process_batch_file(
    content: bytes,
    filename: str,
    output_format: str = "auto",
) -> BatchFileResult:
    lower = (filename or "").lower()
    delimiter = ";"
    if lower.endswith((".xlsx", ".xlsm")):
        rows = parse_xlsx_table(content)
    elif lower.endswith(".xls"):
        rows = parse_xls_table(content)
    else:
        rows, delimiter = parse_text_table(content)

    enriched, stats = enrich_rows(rows)
    fmt = resolve_output_format(filename, output_format)
    if fmt == "xlsx":
        payload = rows_to_xlsx(enriched)
    elif fmt == "xls":
        payload = rows_to_xls(enriched)
    else:
        payload = rows_to_csv(enriched, delimiter)

    return BatchFileResult(
        content=payload,
        media_type=MEDIA_TYPES[fmt],
        filename=output_filename(fmt),
        output_format=fmt,
        stats=stats,
    )
