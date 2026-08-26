from io import BytesIO
from pathlib import Path

import openpyxl

from app.engine.lookups import TEMPLATE_COL_COUNT, TEMPLATE_COLUMNS
from app.engine.normalize import has_value, normalize_text

NUMERIC_FIELDS = frozenset({"impressions", "amount"})
TEXT_NUMBER_FORMAT = "@"
IMPRESSIONS_NUMBER_FORMAT = "0"
AMOUNT_NUMBER_FORMAT = "General"


def load_template_headers(template_path: Path):
    tpl_wb = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    tpl_ws = tpl_wb["Sheet1"]
    headers = [tpl_ws.cell(1, c).value for c in range(1, TEMPLATE_COL_COUNT + 1)]
    descriptions = [tpl_ws.cell(2, c).value for c in range(1, TEMPLATE_COL_COUNT + 1)]
    tpl_wb.close()
    return headers, descriptions


def create_output_workbook(headers, descriptions):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for col, value in enumerate(headers, 1):
        if has_value(value):
            ws.cell(1, col, value)
    for col, value in enumerate(descriptions, 1):
        if has_value(value):
            ws.cell(2, col, value)
    return wb, ws


def _apply_cell_format(cell, field_key: str, value) -> None:
    if field_key in NUMERIC_FIELDS:
        cell.value = value
        if field_key == "impressions":
            cell.number_format = IMPRESSIONS_NUMBER_FORMAT
        else:
            cell.number_format = AMOUNT_NUMBER_FORMAT
        return
    cell.value = normalize_text(value) or str(value).strip()
    cell.number_format = TEXT_NUMBER_FORMAT


def write_cell(ws, row, col, field_key: str, value):
    if not has_value(value):
        return
    cell = ws.cell(row, col)
    _apply_cell_format(cell, field_key, value)


def write_records(ws, records, start_row=3):
    for i, record in enumerate(records):
        out_row = start_row + i
        for key, col in TEMPLATE_COLUMNS.items():
            write_cell(ws, out_row, col, key, record.get(key))


def workbook_to_bytes(wb) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
