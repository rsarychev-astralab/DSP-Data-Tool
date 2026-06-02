"""Чтение старых .xls через xlrd (iter_rows как у openpyxl)."""

from __future__ import annotations

from pathlib import Path
from typing import BinaryIO, Iterator

import xlrd


class XlsSheet:
    def __init__(self, sheet):
        self._sheet = sheet

    def iter_rows(
        self,
        min_row: int = 1,
        max_row: int | None = None,
        min_col: int = 1,
        max_col: int | None = None,
        values_only: bool = True,
    ):
        start_r = max(min_row - 1, 0)
        end_r = self._sheet.nrows if max_row is None else min(max_row, self._sheet.nrows)
        start_c = max(min_col - 1, 0)
        end_c = self._sheet.ncols if max_col is None else min(max_col, self._sheet.ncols)
        for r in range(start_r, end_r):
            yield tuple(self._sheet.cell_value(r, c) for c in range(start_c, end_c))


class XlsWorkbook:
    def __init__(self, path: Path | None = None, content: bytes | None = None):
        if content is not None:
            self._book = xlrd.open_workbook(file_contents=content)
        elif path is not None:
            self._book = xlrd.open_workbook(str(path))
        else:
            raise ValueError("path or content required")

    @property
    def sheetnames(self) -> list[str]:
        return self._book.sheet_names()

    def __getitem__(self, name: str) -> XlsSheet:
        return XlsSheet(self._book.sheet_by_name(name))

    def close(self):
        pass


def is_xls_path(path: Path | str | None) -> bool:
    return bool(path) and str(path).lower().endswith(".xls")


def open_source_workbook(source: BinaryIO | Path, *, filename: str | None = None):
    """Возвращает (workbook, is_xls). Для .xls — XlsWorkbook, иначе openpyxl."""
    import openpyxl
    from io import BytesIO

    if isinstance(source, Path):
        if is_xls_path(source):
            return XlsWorkbook(path=source), True
        return openpyxl.load_workbook(source, read_only=True, data_only=True), False

    data = source.read()
    name = (filename or "").lower()
    if name.endswith(".xls") and not name.endswith(".xlsx"):
        return XlsWorkbook(content=data), True
    return openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True), False
