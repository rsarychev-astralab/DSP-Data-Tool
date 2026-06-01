from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, BinaryIO

import openpyxl

from app.config import TEMPLATE_PATH
from app.engine.normalize import has_value, normalize_field
from app.engine.template import (
    create_output_workbook,
    load_template_headers,
    workbook_to_bytes,
    write_records,
)
from app.profiles.loader import PartnerProfile


@dataclass
class TransformResult:
    output_bytes: bytes
    rows_written: int
    skipped_empty_rows: int
    filename: str


def _is_empty_marker(val: Any, markers: tuple[str, ...]) -> bool:
    if not markers or val is None:
        return False
    text = str(val).strip().upper()
    return text in {m.upper() for m in markers}


def _sanitize_value(val: Any, profile: PartnerProfile) -> Any:
    if _is_empty_marker(val, profile.empty_markers):
        return None
    return val


def _sanitize_raw(raw: dict, profile: PartnerProfile) -> dict:
    if not profile.empty_markers:
        return raw
    return {key: _sanitize_value(val, profile) for key, val in raw.items()}


def _extract_row_values(row: tuple, profile: PartnerProfile) -> dict:
    raw = {
        key: row[idx] if idx < len(row) else None
        for key, idx in profile.column_map.items()
    }
    return _sanitize_raw(raw, profile)


def _resolve_amount(raw: dict, profile: PartnerProfile) -> Any:
    rules = profile.amount_by_vat
    if not rules:
        return raw.get("amount")
    vat_field = rules.get("vat_field", "vat_included")
    vat_val = raw.get(vat_field)
    vat_norm = normalize_field(vat_field, vat_val)
    col_key = "with_vat" if vat_norm == "yes" else "without_vat"
    col_idx = rules.get(col_key)
    if col_idx is None:
        return raw.get("amount")
    return raw.get(f"_col_{col_idx}") if f"_col_{col_idx}" in raw else None


def build_record(raw: dict, profile: PartnerProfile) -> dict:
    record = {}
    field_keys = set(profile.column_map) | set(profile.constants)
    if profile.amount_by_vat:
        field_keys.add("amount")

    # Inject column values for amount_by_vat resolution
    enriched = dict(raw)

    def set_field(key, value):
        if value is not None and has_value(value):
            record[key] = value

    for key in field_keys:
        if key == "amount" and profile.amount_by_vat:
            val = _resolve_amount(enriched, profile)
        else:
            val = raw.get(key)
            if not has_value(val) and key in profile.constants:
                val = profile.constants[key]
        if has_value(val):
            set_field(key, normalize_field(key, val))

    return record


def transform_source(
    source: BinaryIO | Path,
    profile: PartnerProfile,
    *,
    template_path: Path = TEMPLATE_PATH,
    output_filename: str = "result.xlsx",
) -> TransformResult:
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    headers, descriptions = load_template_headers(template_path)
    wb, ws = create_output_workbook(headers, descriptions)

    if isinstance(source, Path):
        src_wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    else:
        src_wb = openpyxl.load_workbook(BytesIO(source.read()), read_only=True, data_only=True)

    if profile.sheet not in src_wb.sheetnames:
        available = ", ".join(src_wb.sheetnames)
        src_wb.close()
        raise ValueError(f"Sheet {profile.sheet!r} not found. Available: {available}")

    src_ws = src_wb[profile.sheet]
    records = []
    skipped_empty = 0

    for row in src_ws.iter_rows(min_row=profile.data_from_row, values_only=True):
        if all(not has_value(v) for v in row):
            skipped_empty += 1
            continue
        raw = _extract_row_values(row, profile)
        for key, idx in profile.column_map.items():
            raw[f"_col_{idx}"] = _sanitize_value(
                row[idx] if idx < len(row) else None, profile
            )
        if profile.amount_by_vat:
            for rule_key in ("with_vat", "without_vat"):
                idx = profile.amount_by_vat.get(rule_key)
                if idx is not None:
                    raw[f"_col_{idx}"] = _sanitize_value(
                        row[idx] if idx < len(row) else None, profile
                    )
        record = build_record(raw, profile)
        if record:
            records.append(record)

    write_records(ws, records)
    output_bytes = workbook_to_bytes(wb)
    src_wb.close()
    wb.close()

    return TransformResult(
        output_bytes=output_bytes,
        rows_written=len(records),
        skipped_empty_rows=skipped_empty,
        filename=output_filename,
    )
