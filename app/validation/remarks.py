from __future__ import annotations

import re
from io import BytesIO

import openpyxl

from app.validation.validate import ValidationResult

_ROW_RE = re.compile(r"^Строка (\d+):\s*(.+)$", re.DOTALL)


def build_remarks_filename(output_filename: str) -> str:
    if output_filename.lower().endswith(".xlsx"):
        return output_filename[:-5] + "_log.xlsx"
    return output_filename + "_log.xlsx"


def _parse_error(error: str) -> tuple[int | None, str]:
    match = _ROW_RE.match(error.strip())
    if not match:
        return None, error.strip()
    return int(match.group(1)), match.group(2).strip()


def build_validation_remarks_bytes(
    validation: ValidationResult,
    *,
    partner_name: str | None = None,
    source_filename: str | None = None,
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Замечания"

    meta_row = 1
    if partner_name:
        ws.cell(meta_row, 1, "DSP")
        ws.cell(meta_row, 2, partner_name)
        meta_row += 1
    if source_filename:
        ws.cell(meta_row, 1, "Исходный файл")
        ws.cell(meta_row, 2, source_filename)
        meta_row += 1
    if validation.errors:
        ws.cell(meta_row, 1, "Всего замечаний")
        ws.cell(meta_row, 2, len(validation.errors))
        meta_row += 1

    header_row = meta_row
    ws.cell(header_row, 1, "Строка Excel")
    ws.cell(header_row, 2, "Замечание")

    data_row = header_row + 1
    for error in validation.errors:
        row_num, message = _parse_error(error)
        ws.cell(data_row, 1, row_num if row_num is not None else "")
        ws.cell(data_row, 2, message if row_num is not None else error)
        data_row += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 80

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
