import csv
import io
import re
from pathlib import Path

import xlrd
import xlwt
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook

MAX_BATCH_INNS = 2000
MAX_UPLOAD_BYTES = 5 * 1024 * 1024
BATCH_DELAY_SEC = 0.08
JOB_TTL_SEC = 60 * 60
INN_RE = re.compile(r"(?<!\d)\d{10}(?:\d{2})?(?!\d)")
INN_HEADERS = {"inn", "инн", "inn_ul", "инн организации"}
RESULT_FIELDS = [
    "inn",
    "status",
    "org_status",
    "type",
    "name_short",
    "name_full",
    "ogrn",
    "kpp",
    "okved",
    "manager",
    "address",
    "error",
]
RESULT_HEADERS_RU = {
    "inn": "ИНН",
    "status": "Статус проверки",
    "org_status": "Статус организации",
    "type": "Тип",
    "name_short": "Краткое название",
    "name_full": "Полное название",
    "ogrn": "ОГРН",
    "kpp": "КПП",
    "okved": "ОКВЭД",
    "manager": "Руководитель",
    "address": "Адрес",
    "error": "Ошибка",
}
STATUS_LABELS = {
    "ACTIVE": "Действующая",
    "LIQUIDATING": "Ликвидируется",
    "LIQUIDATED": "Ликвидирована",
    "BANKRUPT": "Банкротство",
    "REORGANIZING": "Реорганизация",
}
ALLOWED_INPUT = {".csv", ".txt", ".tsv", ".xls", ".xlsx"}
ALLOWED_OUTPUT = {"csv", "xls", "xlsx"}
INN_10_WEIGHTS = (2, 4, 10, 3, 5, 9, 4, 6, 8)
INN_11_WEIGHTS = (7, 2, 4, 10, 3, 5, 9, 4, 6, 8)
INN_12_WEIGHTS = (3, 7, 2, 4, 10, 3, 5, 9, 4, 6, 8)


def inn_checksum_ok(inn: str) -> bool:
    if not inn.isdigit():
        return False
    if set(inn) == {"0"}:
        return False
    digits = [int(ch) for ch in inn]
    if len(digits) == 10:
        control = sum(d * w for d, w in zip(digits[:9], INN_10_WEIGHTS)) % 11 % 10
        return control == digits[9]
    if len(digits) == 12:
        n11 = sum(d * w for d, w in zip(digits[:10], INN_11_WEIGHTS)) % 11 % 10
        n12 = sum(d * w for d, w in zip(digits[:11], INN_12_WEIGHTS)) % 11 % 10
        return n11 == digits[10] and n12 == digits[11]
    return False


def normalize_inn(value: object | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, float):
        if value.is_integer():
            digits = str(int(value))
        else:
            digits = re.sub(r"\D", "", str(value))
    elif isinstance(value, int):
        digits = str(value)
    else:
        text = str(value).strip()
        if re.fullmatch(r"\d+(\.0+)?([eE][+]?\d+)?", text.replace(" ", "")):
            try:
                as_float = float(text.replace(" ", "").replace(",", "."))
                if as_float.is_integer():
                    digits = str(int(as_float))
                else:
                    digits = re.sub(r"\D", "", text)
            except ValueError:
                digits = re.sub(r"\D", "", text)
        else:
            digits = re.sub(r"\D", "", text)

    if len(digits) in (10, 12):
        return digits
    if digits.isdigit() and 1 <= len(digits) < 10:
        padded = digits.zfill(10)
        if inn_checksum_ok(padded):
            return padded
    if digits.isdigit() and 10 < len(digits) < 12:
        padded = digits.zfill(12)
        if inn_checksum_ok(padded):
            return padded
    return None


def extract_inns_from_text(text: str) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for match in INN_RE.finditer(text):
        inn = match.group(0)
        if inn not in seen:
            seen.add(inn)
            ordered.append(inn)
    return ordered


def unique_inns(inns: list[str]) -> list[str]:
    unique: list[str] = []
    seen: set[str] = set()
    for inn in inns:
        if inn not in seen:
            seen.add(inn)
            unique.append(inn)

    if not unique:
        raise HTTPException(
            status_code=400,
            detail="В файле не найдено ни одного ИНН (10 или 12 цифр)",
        )
    if len(unique) > MAX_BATCH_INNS:
        raise HTTPException(
            status_code=400,
            detail=f"Слишком много ИНН: {len(unique)}. Лимит за один раз: {MAX_BATCH_INNS}",
        )
    return unique


def inns_from_table_rows(rows: list[list[object]]) -> list[str]:
    if not rows:
        return []

    header = [str(cell or "").strip().lower() for cell in rows[0]]
    inn_col = None
    for idx, name in enumerate(header):
        if name in INN_HEADERS:
            inn_col = idx
            break

    inns: list[str] = []
    start = 1 if inn_col is not None else 0
    for row in rows[start:]:
        if not row:
            continue
        if inn_col is not None and inn_col < len(row):
            inn = normalize_inn(row[inn_col])
            if inn:
                inns.append(inn)
            continue
        found_in_row = False
        for cell in row:
            inn = normalize_inn(cell)
            if inn:
                inns.append(inn)
                found_in_row = True
                break
        if not found_in_row:
            text = ";".join("" if cell is None else str(cell) for cell in row)
            inns.extend(extract_inns_from_text(text))
    return inns


def parse_text_file(content: bytes, filename: str) -> list[str]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("cp1251", errors="replace")

    text = text.strip()
    if not text:
        raise HTTPException(status_code=400, detail="Файл пустой")

    lower_name = (filename or "").lower()
    if lower_name.endswith((".csv", ".tsv", ".txt")) or ";" in text or "," in text or "\t" in text:
        sample = text[:4096]
        delimiter = ";"
        if sample.count("\t") > sample.count(";") and sample.count("\t") > sample.count(","):
            delimiter = "\t"
        elif sample.count(",") > sample.count(";"):
            delimiter = ","
        rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
        inns = inns_from_table_rows(rows)
    else:
        inns = extract_inns_from_text(text)

    return unique_inns(inns)


def parse_xlsx(content: bytes) -> list[str]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Не удалось прочитать XLSX: {exc}") from exc

    sheet = workbook.active
    rows = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    return unique_inns(inns_from_table_rows(rows))


def parse_xls(content: bytes) -> list[str]:
    try:
        book = xlrd.open_workbook(file_contents=content)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Не удалось прочитать XLS: {exc}") from exc

    sheet = book.sheet_by_index(0)
    rows = [sheet.row_values(row_idx) for row_idx in range(sheet.nrows)]
    return unique_inns(inns_from_table_rows(rows))


def parse_uploaded_inns(content: bytes, filename: str) -> list[str]:
    lower = (filename or "").lower()
    if lower.endswith(".xlsx"):
        return parse_xlsx(content)
    if lower.endswith(".xls"):
        return parse_xls(content)
    return parse_text_file(content, filename)


def party_to_row(inn: str, party: dict | None, error: str = "") -> dict[str, str]:
    empty = {
        "inn": inn,
        "status": "ошибка" if error else "не найдено",
        "org_status": "",
        "type": "",
        "name_short": "",
        "name_full": "",
        "ogrn": "",
        "kpp": "",
        "okved": "",
        "manager": "",
        "address": "",
        "error": error,
    }
    if error or not party:
        return empty

    data = party.get("data") or {}
    name = data.get("name") or {}
    state = data.get("state") or {}
    management = data.get("management") or {}
    address = data.get("address") or {}
    org_status = state.get("status") or ""
    type_label = "ИП" if data.get("type") == "INDIVIDUAL" else "ЮЛ"

    return {
        "inn": data.get("inn") or inn,
        "status": "найдено",
        "org_status": STATUS_LABELS.get(org_status, org_status),
        "type": type_label,
        "name_short": name.get("short_with_opf") or party.get("value") or "",
        "name_full": name.get("full_with_opf") or "",
        "ogrn": data.get("ogrn") or "",
        "kpp": data.get("kpp") or "",
        "okved": data.get("okved") or "",
        "manager": management.get("name") or "",
        "address": address.get("unrestricted_value") or address.get("value") or "",
        "error": "",
    }


def rows_to_csv(rows: list[dict[str, str]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.DictWriter(
        buffer,
        fieldnames=RESULT_FIELDS,
        delimiter=";",
        lineterminator="\n",
    )
    writer.writerow({key: RESULT_HEADERS_RU[key] for key in RESULT_FIELDS})
    for row in rows:
        writer.writerow({key: row.get(key, "") for key in RESULT_FIELDS})
    return ("\ufeff" + buffer.getvalue()).encode("utf-8")


def rows_to_xlsx(rows: list[dict[str, str]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Результат"
    sheet.append([RESULT_HEADERS_RU[key] for key in RESULT_FIELDS])
    for row in rows:
        sheet.append([row.get(key, "") for key in RESULT_FIELDS])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def rows_to_xls(rows: list[dict[str, str]]) -> bytes:
    workbook = xlwt.Workbook(encoding="utf-8")
    sheet = workbook.add_sheet("Результат")
    for col, key in enumerate(RESULT_FIELDS):
        sheet.write(0, col, RESULT_HEADERS_RU[key])
    for row_idx, row in enumerate(rows, start=1):
        for col, key in enumerate(RESULT_FIELDS):
            sheet.write(row_idx, col, row.get(key, ""))
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def default_output_format(filename: str) -> str:
    lower = (filename or "").lower()
    if lower.endswith(".xlsx"):
        return "xlsx"
    if lower.endswith(".xls"):
        return "xls"
    return "csv"


def build_result_file(rows: list[dict[str, str]], output_format: str) -> tuple[bytes, str, str]:
    if output_format == "xlsx":
        return (
            rows_to_xlsx(rows),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "dadata_result.xlsx",
        )
    if output_format == "xls":
        return (
            rows_to_xls(rows),
            "application/vnd.ms-excel",
            "dadata_result.xls",
        )
    return (
        rows_to_csv(rows),
        "text/csv; charset=utf-8",
        "dadata_result.csv",
    )


def suffix_allowed(filename: str) -> bool:
    return Path(filename).suffix.lower() in ALLOWED_INPUT
