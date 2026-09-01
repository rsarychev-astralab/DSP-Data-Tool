from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

from app.config import TEMPLATE_PATH
from app.engine.transform import build_record, transform_source
from app.profiles.loader import load_profile


def test_build_record_empty_activity_stays_empty():
    profile = load_profile("genius_desk")
    record = build_record(
        {
            "erid": "abc",
            "contract_no": "1",
            "contract_subject": "org-distribution",
            "activity_type": None,
        },
        profile,
    )
    assert "activity_type" not in record


def test_build_record_activity_from_source_is_normalized():
    profile = load_profile("genius_desk")
    record = build_record(
        {
            "erid": "abc",
            "activity_type": "distribution",
        },
        profile,
    )
    assert record["activity_type"] == "Distribution"


def test_buzzoola_maps_address_columns():
    profile = load_profile("buzzoola")
    assert profile.column_map["customer_address"] == 26
    assert profile.column_map["contractor_address"] == 21
    record = build_record(
        {
            "erid": "2VfnxxpZ3Yd",
            "customer_address": "129085, г. Москва, ул. Годовикова, д. 9, стр. 10.",
            "contractor_address": "197110, город Санкт-Петербург, Константиновский пр-кт, д. 11",
        },
        profile,
    )
    assert record["customer_address"] == "129085, г. Москва, ул. Годовикова, д. 9, стр. 10."
    assert record["contractor_address"].startswith("197110")


def test_transform_buzzoola_writes_addresses():
    source = Path("Исходные данные/buzzoola.xlsx")
    if not source.exists():
        pytest.skip("buzzoola sample not present")
    profile = load_profile("buzzoola")
    result = transform_source(
        source,
        profile,
        template_path=TEMPLATE_PATH,
        output_filename="buzzoola.xlsx",
        source_filename=source.name,
    )
    assert result.rows_written > 0
    with_customer = sum(1 for rec in result.records if rec.get("customer_address"))
    with_contractor = sum(1 for rec in result.records if rec.get("contractor_address"))
    assert with_customer > 0
    assert with_contractor > 0


def test_buzzoola_amount_by_vat_without_vat_uses_without_vat_column():
    profile = load_profile("buzzoola")
    record = build_record(
        {
            "erid": "abc",
            "vat_included": "без НДС",
            "_col_7": 10.5,
            "_col_10": 12.6,
        },
        profile,
    )
    assert record["vat_included"] == "no"
    assert record["amount"] == 10.5


def test_buzzoola_amount_by_vat_skips_when_vat_empty():
    profile = load_profile("buzzoola")
    record = build_record(
        {
            "erid": "abc",
            "vat_included": None,
            "_col_7": 10.5,
            "_col_10": 12.6,
        },
        profile,
    )
    assert "amount" not in record
    assert "vat_included" not in record


def test_transform_reports_skip_reasons():
    profile = load_profile("genius_desk")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лист2"
    headers = [""] * 20
    headers[0] = "Номер_Договора"
    headers[5] = "erid"
    headers[7] = "Тип_Рекламодателя"
    headers[13] = "imps"
    for col, val in enumerate(headers, 1):
        ws.cell(1, col, val)
    ws.cell(2, 1, "DOG-1")
    ws.cell(2, 6, "")
    ws.cell(2, 13, "Agency")
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    result = transform_source(
        buf,
        profile,
        template_path=TEMPLATE_PATH,
        output_filename="t.xlsx",
        source_filename="t.xlsx",
    )
    assert result.skipped_empty_rows >= 1
    assert result.skipped_reasons
    assert any("нет ERID" in r or "пустая" in r for r in result.skipped_reasons)
