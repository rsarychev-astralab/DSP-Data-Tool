from io import BytesIO

import openpyxl

from app.engine.template import (
    AMOUNT_NUMBER_FORMAT,
    IMPRESSIONS_NUMBER_FORMAT,
    TEXT_NUMBER_FORMAT,
    create_output_workbook,
    write_cell,
    write_records,
)


def test_write_cell_text_format_for_inn():
    n = 21
    wb, ws = create_output_workbook([""] * n, [""] * n)
    write_cell(ws, 3, 10, "customer_inn", "7700123456789")
    cell = ws.cell(3, 10)
    assert cell.number_format == TEXT_NUMBER_FORMAT
    assert cell.value == "7700123456789"
    assert isinstance(cell.value, str)


def test_write_cell_numeric_format_for_impressions_and_amount():
    n = 21
    wb, ws = create_output_workbook([""] * n, [""] * n)
    write_cell(ws, 3, 20, "impressions", 1200)
    write_cell(ws, 3, 21, "amount", 1500.5)
    assert ws.cell(3, 20).number_format == IMPRESSIONS_NUMBER_FORMAT
    assert ws.cell(3, 20).value == 1200
    assert ws.cell(3, 21).number_format == AMOUNT_NUMBER_FORMAT
    assert ws.cell(3, 21).value == 1500.5


def test_write_records_mixed_formats():
    n = 21
    wb, ws = create_output_workbook([""] * n, [""] * n)
    write_records(ws, [{
        "erid": "abc",
        "contract_no": "001",
        "impressions": 10,
        "amount": 99.9,
    }])
    assert ws.cell(3, 1).number_format == TEXT_NUMBER_FORMAT
    assert ws.cell(3, 2).number_format == TEXT_NUMBER_FORMAT
    assert ws.cell(3, 20).number_format == IMPRESSIONS_NUMBER_FORMAT
    assert ws.cell(3, 21).number_format == AMOUNT_NUMBER_FORMAT


def test_write_records_skips_empty_addresses():
    n = 21
    wb, ws = create_output_workbook([""] * n, [""] * n)
    write_records(ws, [{
        "erid": "abc",
        "customer_name": "Customer LLC",
        "contractor_name": "Contractor LLC",
        "customer_address": None,
        "contractor_address": "   ",
    }])
    assert ws.cell(3, 9).value is None
    assert ws.cell(3, 15).value is None


def test_write_records_writes_addresses_when_present():
    n = 21
    wb, ws = create_output_workbook([""] * n, [""] * n)
    write_records(ws, [{
        "erid": "abc",
        "customer_address": "г. Москва, ул. Тверская, 1",
        "contractor_address": "Legal address",
    }])
    assert ws.cell(3, 9).value == "г. Москва, ул. Тверская, 1"
    assert ws.cell(3, 15).value == "Legal address"


def test_output_workbook_bytes_preserve_text_format():
    from app.engine.template import workbook_to_bytes

    n = 21
    wb, ws = create_output_workbook([""] * n, [""] * n)
    write_cell(ws, 3, 1, "erid", "Lq7abc123")
    write_cell(ws, 3, 20, "impressions", 5)
    data = workbook_to_bytes(wb)
    out = openpyxl.load_workbook(BytesIO(data), data_only=False)
    cell = out.active.cell(3, 1)
    assert cell.number_format == TEXT_NUMBER_FORMAT
    assert out.active.cell(3, 20).number_format == IMPRESSIONS_NUMBER_FORMAT
    out.close()


def test_project_template_headers_include_addresses():
    from app.config import TEMPLATE_PATH
    from app.engine.lookups import TEMPLATE_HEADERS
    from app.engine.template import load_template_headers

    if not TEMPLATE_PATH.exists():
        return
    headers, descriptions = load_template_headers(TEMPLATE_PATH)
    assert tuple(headers) == TEMPLATE_HEADERS
    assert headers[8] == "Адрес заказчика"
    assert headers[14] == "Адрес исполнителя"
    assert descriptions[8]
    assert descriptions[14]
