from datetime import datetime
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

from app.matching.keys import build_match_key
from app.matching.match import (
    OUTPUT_HEADERS,
    OUTPUT_SHEET_IDS,
    OUTPUT_SHEET_UPLOAD,
    UPLOAD_DETAIL_COLUMNS,
    match_workbook_bytes,
)
from app.validation.validate import COL_FIELD, HEADERS


def _idx(field: str) -> int:
    return COL_FIELD.index(field)


def _minimal_upload_bytes(**fields) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, h in enumerate(HEADERS, 1):
        ws.cell(1, col, h)
    row = [""] * len(COL_FIELD)
    row[_idx("erid")] = fields.get("erid", "erid-1")
    row[_idx("contract_no")] = fields.get("contract_no", "dog-1")
    row[_idx("contract_date")] = fields.get("contract_date", "2025-01-01")
    row[_idx("contract_type")] = fields.get("contract_type", "Original")
    row[_idx("contract_subject")] = fields.get("contract_subject", "Distribution")
    row[_idx("activity_type")] = fields.get("activity_type", "")
    row[_idx("customer_inn")] = fields.get("customer_inn", "7700000002")
    row[_idx("contractor_inn")] = fields.get("contractor_inn", "7700000001")
    for col, val in enumerate(row, 1):
        if val != "":
            ws.cell(3, col, val)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _minimal_db_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OriginalContract"
    headers = [
        "Номер", "Дата", "Ответственный", "Исполнитель", "ИНН исполнителя",
        "Заказчик", "ИНН заказчика", "Тип договора", "Сведения о предмете договора",
        "Вид деятельности", "Цена", "Включая НДС", "AdX", "Неактуален",
        "ID в OTM", "ID в OZON", "ID в VK",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    ws.append([
        "DOG-42", "2025-01-01", "resp",
        "ООО «Исполнитель»", "7700000001",
        "ООО «Заказчик»", "7700000002",
        "Оказание услуг", "Распространение рекламы", "NA",
        100, "Нет", "Нет", "Нет", 999, 1001, 1002,
    ])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_match_rejects_raw_partner_file(tmp_path, monkeypatch):
    db_path = tmp_path / "База договоров.xlsx"
    db_path.write_bytes(_minimal_db_bytes())
    monkeypatch.setattr("app.matching.match.resolve_contract_attrs_path", lambda: db_path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1, "Erid")
    ws.cell(1, 2, "Partner column")
    ws.cell(3, 1, "token-1")
    buf = BytesIO()
    wb.save(buf)

    with pytest.raises(ValueError, match="после преобразования"):
        match_workbook_bytes(buf.getvalue())


def test_match_requires_contract_attrs(monkeypatch):
    monkeypatch.setattr("app.matching.match.resolve_contract_attrs_path", lambda: None)
    data = _minimal_upload_bytes()
    with pytest.raises(FileNotFoundError, match="атрибутов договоров"):
        match_workbook_bytes(data)


def test_match_finds_contract_and_outputs_ids(tmp_path, monkeypatch):
    db_path = tmp_path / "База договоров.xlsx"
    db_path.write_bytes(_minimal_db_bytes())
    monkeypatch.setattr("app.matching.match.resolve_contract_attrs_path", lambda: db_path)

    upload = _minimal_upload_bytes(
        contract_no="DOG-42",
        contract_date="2025-01-01",
        contract_type="Original",
        contract_subject="Distribution",
        activity_type="None",
        customer_inn="7700000002",
        contractor_inn="7700000001",
    )
    result = match_workbook_bytes(upload, original_filename="out.xlsx")
    assert result.rows_total == 1
    assert result.rows_matched == 1
    assert result.rows_unmatched == 0

    out = openpyxl.load_workbook(BytesIO(result.output_bytes), data_only=True)
    assert out.sheetnames == [OUTPUT_SHEET_IDS, OUTPUT_SHEET_UPLOAD]
    ws = out[OUTPUT_SHEET_IDS]
    assert [ws.cell(1, c).value for c in range(1, 5)] == OUTPUT_HEADERS
    assert ws.cell(3, 1).value == "erid-1"
    assert ws.cell(3, 2).value == "999"
    assert ws.cell(3, 3).value == "1001"
    assert ws.cell(3, 4).value == "1002"
    ws2 = out[OUTPUT_SHEET_UPLOAD]
    assert ws2.cell(1, 1).value == "ERID"
    assert ws2.cell(1, 7).value == "Тип заказчика"
    assert ws2.cell(1, len(UPLOAD_DETAIL_COLUMNS) + 1).value == "ID в OTM"
    assert ws2.cell(3, 2).value == "DOG-42"
    assert ws2.cell(3, len(UPLOAD_DETAIL_COLUMNS) + 1).value == "999"
    out.close()


def test_match_by_contract_applies_to_all_erid_rows(tmp_path, monkeypatch):
    """Один договор в базе — все строки с тем же номером получают ID, даже если дата в строке другая."""
    db_path = tmp_path / "База договоров.xlsx"
    db_path.write_bytes(_minimal_db_bytes())
    monkeypatch.setattr("app.matching.match.resolve_contract_attrs_path", lambda: db_path)

    wb = openpyxl.Workbook()
    ws = wb.active
    for col, h in enumerate(HEADERS, 1):
        ws.cell(1, col, h)
    good = [""] * len(COL_FIELD)
    good[_idx("erid")] = "erid-a"
    good[_idx("contract_no")] = "DOG-42"
    good[_idx("contract_date")] = "2025-01-01"
    good[_idx("contract_type")] = "Original"
    good[_idx("contract_subject")] = "Distribution"
    good[_idx("activity_type")] = "None"
    good[_idx("customer_inn")] = "7700000002"
    good[_idx("contractor_inn")] = "7700000001"
    good[_idx("impressions")] = 100
    good[_idx("amount")] = 10
    bad_date = list(good)
    bad_date[_idx("erid")] = "erid-b"
    bad_date[_idx("contract_date")] = "2099-12-31"
    bad_date[_idx("impressions")] = 200
    bad_date[_idx("amount")] = 20
    for col, val in enumerate(good, 1):
        ws.cell(3, col, val)
    for col, val in enumerate(bad_date, 1):
        ws.cell(4, col, val)
    buf = BytesIO()
    wb.save(buf)

    result = match_workbook_bytes(buf.getvalue())
    assert result.contracts_total == 1
    assert result.contracts_matched == 1
    assert result.rows_matched == 2
    out = openpyxl.load_workbook(BytesIO(result.output_bytes), data_only=True)
    assert out[OUTPUT_SHEET_IDS].cell(3, 2).value == "999"
    assert out[OUTPUT_SHEET_UPLOAD].cell(4, len(UPLOAD_DETAIL_COLUMNS) + 1).value == "999"
    out.close()


def test_match_unmatched_row(tmp_path, monkeypatch):
    db_path = tmp_path / "База договоров.xlsx"
    db_path.write_bytes(_minimal_db_bytes())
    monkeypatch.setattr("app.matching.match.resolve_contract_attrs_path", lambda: db_path)

    upload = _minimal_upload_bytes(contract_no="UNKNOWN-999")
    result = match_workbook_bytes(upload)
    assert result.rows_matched == 0
    assert result.rows_unmatched == 1


def test_match_real_database_sample_row():
    db = Path("Справка/База договоров.xlsx")
    if not db.exists():
        pytest.skip("no contract database")
    upload = _minimal_upload_bytes(
        erid="test-erid",
        contract_no="TEST",
        contract_date="2022-08-03",
        contract_type="Original",
        contract_subject="Distribution",
        activity_type="None",
        customer_inn="7713085659",
        contractor_inn="7702848355",
    )
    result = match_workbook_bytes(upload)
    assert result.rows_matched == 1, f"unmatched={result.rows_unmatched}"
    out = openpyxl.load_workbook(BytesIO(result.output_bytes), data_only=True)
    assert out.active.cell(3, 2).value == "317"
    out.close()


def test_build_match_key_db_upload_symmetry():
    db_key = build_match_key(
        contract_no="TEST - 1991",
        contract_date=datetime(2022, 8, 3),
        customer_inn="7713085659",
        contractor_inn="7702848355",
        contract_type="Оказание услуг",
        contract_subject="Распространение рекламы",
        from_db=True,
    )
    up_key = build_match_key(
        contract_no="TEST",
        contract_date="2022-08-03",
        customer_inn="7713085659",
        contractor_inn=7702848355,
        contract_type="Original",
        contract_subject="Distribution",
        from_db=False,
    )
    assert db_key == up_key
