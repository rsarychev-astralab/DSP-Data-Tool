import io

import pytest
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook

from app.dadata.core import (
    MAX_BATCH_INNS,
    build_result_file,
    inn_checksum_ok,
    inns_from_table_rows,
    normalize_inn,
    parse_text_file,
    parse_uploaded_inns,
    parse_xlsx,
    party_to_row,
    unique_inns,
)


def test_inn_checksum_valid_ul():
    assert inn_checksum_ok("7707083893") is True


def test_inn_checksum_valid_ip():
    assert inn_checksum_ok("500100732259") is True


def test_inn_checksum_rejects_all_zeros():
    assert inn_checksum_ok("0000000000") is False
    assert inn_checksum_ok("000000000000") is False


def test_inn_checksum_rejects_wrong_length_and_digits():
    assert inn_checksum_ok("1234567890") is False
    assert inn_checksum_ok("770708389") is False
    assert inn_checksum_ok("77070838931") is False
    assert inn_checksum_ok("7707083893a") is False


def test_normalize_inn_from_float_and_int():
    assert normalize_inn(7707083893) == "7707083893"
    assert normalize_inn(7707083893.0) == "7707083893"
    assert normalize_inn("7707083893.0") == "7707083893"


def test_inns_from_table_rows_uses_inn_header():
    rows = [
        ["name", "ИНН", "other"],
        ["Сбер", "7707083893", "x"],
        ["ИП", "500100732259", "y"],
        ["пусто", "", "z"],
    ]
    assert inns_from_table_rows(rows) == ["7707083893", "500100732259"]


def test_inns_from_table_rows_scans_cells_without_header():
    rows = [
        ["7707083893", "комментарий"],
        ["не инн", "500100732259"],
    ]
    assert inns_from_table_rows(rows) == ["7707083893", "500100732259"]


def test_unique_inns_dedupes_and_rejects_empty():
    assert unique_inns(["7707083893", "7707083893", "500100732259"]) == [
        "7707083893",
        "500100732259",
    ]
    with pytest.raises(HTTPException) as exc:
        unique_inns([])
    assert exc.value.status_code == 400


def test_unique_inns_rejects_over_limit():
    inns = [f"{i:010d}" for i in range(MAX_BATCH_INNS + 1)]
    with pytest.raises(HTTPException) as exc:
        unique_inns(inns)
    assert exc.value.status_code == 400


def test_parse_csv_with_header():
    content = "ИНН\n7707083893\n500100732259\n".encode("utf-8")
    assert parse_text_file(content, "inns.csv") == ["7707083893", "500100732259"]


def test_parse_xlsx_with_header():
    wb = Workbook()
    ws = wb.active
    ws.append(["ИНН"])
    ws.append(["7707083893"])
    ws.append(["500100732259"])
    buf = io.BytesIO()
    wb.save(buf)
    assert parse_xlsx(buf.getvalue()) == ["7707083893", "500100732259"]


def test_parse_uploaded_inns_empty_csv_errors():
    with pytest.raises(HTTPException) as exc:
        parse_uploaded_inns("не числа".encode("utf-8"), "inns.csv")
    assert exc.value.status_code == 400


def test_party_to_row_found_and_missing():
    missing = party_to_row("7707083893", None)
    assert missing["status"] == "не найдено"
    assert missing["error"] == ""

    err = party_to_row("7707083893", None, error="лимит")
    assert err["status"] == "ошибка"
    assert err["error"] == "лимит"

    party = {
        "value": "ПАО Сбербанк",
        "data": {
            "inn": "7707083893",
            "type": "LEGAL",
            "ogrn": "1027700132195",
            "kpp": "773601001",
            "okved": "64.19",
            "name": {
                "short_with_opf": "ПАО Сбербанк",
                "full_with_opf": "ПАО Сбербанк полное",
            },
            "state": {"status": "ACTIVE"},
            "management": {"name": "Иванов И.И."},
            "address": {"unrestricted_value": "Москва"},
        },
    }
    row = party_to_row("7707083893", party)
    assert row["status"] == "найдено"
    assert row["org_status"] == "Действующая"
    assert row["type"] == "ЮЛ"
    assert row["name_short"] == "ПАО Сбербанк"
    assert row["address"] == "Москва"


def test_build_result_file_xlsx_and_csv():
    rows = [party_to_row("7707083893", None, error="тест")]
    payload, media, name = build_result_file(rows, "xlsx")
    assert name == "dadata_result.xlsx"
    assert "spreadsheet" in media
    wb = load_workbook(io.BytesIO(payload))
    ws = wb.active
    assert ws.cell(1, 1).value == "ИНН"
    assert ws.cell(2, 1).value == "7707083893"

    csv_payload, csv_media, csv_name = build_result_file(rows, "csv")
    assert csv_name == "dadata_result.csv"
    assert csv_payload.startswith(b"\xef\xbb\xbf")
    assert "ИНН" in csv_payload.decode("utf-8-sig")
    assert csv_media.startswith("text/csv")
