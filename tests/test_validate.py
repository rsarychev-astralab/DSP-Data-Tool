from app.profiles.loader import load_profile
from app.validation.validate import COL_FIELD, validate_row, validate_workbook_bytes


def _idx(field: str) -> int:
    return COL_FIELD.index(field)


def _empty_row():
    return [""] * len(COL_FIELD)


def _full_row(**overrides):
    row = _empty_row()
    defaults = {
        _idx("erid"): "erid-1",
        _idx("contract_no"): "dog-1",
        _idx("contract_date"): "2025-01-01",
        _idx("contract_type"): "Original",
        _idx("contract_subject"): "Distribution",
        _idx("activity_type"): "Distribution",
        _idx("customer_type"): "LegalPerson",
        _idx("customer_name"): "Customer LLC",
        _idx("customer_inn"): "7700000000",
        _idx("contractor_type"): "LegalPerson",
        _idx("contractor_name"): "Contractor LLC",
        _idx("contractor_inn"): "7700000001",
        _idx("vat_included"): "yes",
        _idx("impressions"): "100",
        _idx("amount"): "1000",
    }
    defaults.update(overrides)
    for idx, val in defaults.items():
        row[idx] = val
    return row


def test_validate_row_requires_erid():
    errors = validate_row(3, _empty_row(), None)
    assert any("ERID" in e for e in errors)
    assert all(e.startswith("Строка 3:") for e in errors)


def test_validate_workbook_collects_row_numbers():
    from io import BytesIO

    import openpyxl  # noqa: PLC0415

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["h1", "h2"])
    ws.append(["h1", "h2"])
    ws.append(_empty_row())
    ws.append(["erid"] + [""] * (len(COL_FIELD) - 1))
    buf = BytesIO()
    wb.save(buf)
    validation = validate_workbook_bytes(buf.getvalue(), None)
    assert 4 in validation.row_numbers
    assert any("Строка 4:" in e for e in validation.errors)


def test_adriver_allows_empty_reg_no_for_ru():
    profile = load_profile("adriver")
    row = _full_row()
    errors = validate_row(3, row, profile)
    assert not any("Рег.номер" in e for e in errors)
    assert not any("ОКСМ" in e for e in errors)
    assert not any("НДС" in e for e in errors)
    assert not any("Показы" in e for e in errors)


def test_validate_row_accepts_normalized_contract_type():
    row = _full_row()
    errors = validate_row(3, row, None)
    assert not any("недопустимое" in e and "Тип договора" in e for e in errors)


def test_validate_row_rejects_bad_contract_type():
    row = _full_row()
    row[_idx("contract_type")] = "NotARealType"
    errors = validate_row(3, row, None)
    assert any("Тип договора" in e for e in errors)


def test_validate_row_reports_empty_field_by_name():
    row = _full_row()
    row[_idx("activity_type")] = ""
    errors = validate_row(3, row, None)
    assert any("Строка 3: поле «Вид деятельности» пустое" in e for e in errors)


def test_validate_row_allows_empty_addresses():
    row = _full_row()
    row[_idx("customer_address")] = ""
    row[_idx("contractor_address")] = ""
    errors = validate_row(3, row, None)
    assert not any("Адрес" in e for e in errors)


def test_ru_party_requires_inn():
    profile = load_profile("adriver")
    row = _full_row()
    row[_idx("customer_inn")] = ""
    errors = validate_row(3, row, profile)
    assert any("ИНН" in e and "заказчика" in e for e in errors)


def test_adriver_transform_preserves_activity_from_source():
    from io import BytesIO
    from collections import Counter
    from pathlib import Path

    import openpyxl

    from app.config import TEMPLATE_PATH
    from app.engine.transform import transform_source
    from app.validation.validate import validate_workbook_bytes

    source = Path("Исходные данные/adriver.xlsx")
    if not source.exists():
        return
    profile = load_profile("adriver")
    act_col = profile.column_map["activity_type"]
    src_counts = Counter()
    wb_src = openpyxl.load_workbook(source, read_only=True, data_only=True)
    ws_src = wb_src[profile.sheet]
    for row in ws_src.iter_rows(min_row=profile.data_from_row, values_only=True):
        if not row:
            continue
        val = row[act_col] if act_col < len(row) else None
        src_counts["filled" if val and str(val).strip() else "empty"] += 1
    wb_src.close()

    with source.open("rb") as f:
        result = transform_source(
            BytesIO(f.read()), profile, template_path=TEMPLATE_PATH
        )
    wb_out = openpyxl.load_workbook(BytesIO(result.output_bytes), read_only=True, data_only=True)
    ws_out = wb_out.active
    out_counts = Counter()
    for row in ws_out.iter_rows(min_row=3, min_col=6, max_col=6, values_only=True):
        val = row[0] if row else None
        out_counts["filled" if val and str(val).strip() else "empty"] += 1
    wb_out.close()

    validation = validate_workbook_bytes(result.output_bytes, profile)
    assert not any(
        "недопустимое" in e and "Вид деятельности" in e for e in validation.errors
    )
    assert out_counts["filled"] > 0
    assert out_counts["empty"] > 0
    assert src_counts["empty"] > 0


def test_validate_records_matches_workbook_validation():
    from io import BytesIO

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["h1", "h2"])
    ws.append(["h1", "h2"])
    ws.append(["erid"] + [""] * (len(COL_FIELD) - 1))
    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    from app.validation.validate import validate_records

    records = [{"erid": "erid"}]
    by_records = validate_records(records, None)
    by_workbook = validate_workbook_bytes(data, None)
    assert by_records.row_numbers == by_workbook.row_numbers
    assert len(by_records.errors) == len(by_workbook.errors)
