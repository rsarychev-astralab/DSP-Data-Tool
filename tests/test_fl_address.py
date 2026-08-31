import io

import pytest
from fastapi import HTTPException
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.dadata.core import INN_11_WEIGHTS, INN_12_WEIGHTS, inn_checksum_ok
from app.fl_address.batch import enrich_rows, process_batch_file
from app.fl_address.core import inn_digits, lookup_fl_address, resolve_fl_address
from app.fl_address.regions import REGIONS, address_for_region
from app.main import app

client = TestClient(app)


def make_inn12(body10: str) -> str:
    digits = [int(ch) for ch in body10]
    n11 = sum(d * w for d, w in zip(digits, INN_11_WEIGHTS)) % 11 % 10
    digits.append(n11)
    n12 = sum(d * w for d, w in zip(digits, INN_12_WEIGHTS)) % 11 % 10
    digits.append(n12)
    inn = "".join(str(d) for d in digits)
    assert inn_checksum_ok(inn)
    return inn


def test_lookup_known_ip_inn_moscow_oblast():
    result = lookup_fl_address("500100732259")
    assert result["inn"] == "500100732259"
    assert result["region_code"] == "50"
    assert result["region"] == "Московская область"
    assert result["address"] == "Российская Федерация, Московская область"


def test_lookup_moscow_and_spb_aliases():
    moscow = lookup_fl_address(make_inn12("7700000001"))
    assert moscow["region"] == "г. Москва"
    assert moscow["address"] == "Российская Федерация, г. Москва"

    spb = lookup_fl_address(make_inn12("9800000001"))
    assert spb["region"] == "г. Санкт-Петербург"


def test_lookup_rejects_ul_inn():
    with pytest.raises(HTTPException) as exc_info:
        lookup_fl_address("7707083893")
    assert exc_info.value.status_code == 400
    assert "юрлицо" in exc_info.value.detail


def test_lookup_rejects_bad_checksum():
    with pytest.raises(HTTPException) as exc_info:
        lookup_fl_address("500100732250")
    assert exc_info.value.status_code == 400
    assert "контрольная сумма" in exc_info.value.detail


def test_lookup_accepts_inn_with_spaces():
    result = lookup_fl_address("5001 0073 2259")
    assert result["region_code"] == "50"


def test_every_region_code_has_address():
    for code, region in REGIONS.items():
        assert address_for_region(region).startswith("Российская Федерация, ")
        assert code.isdigit() and len(code) == 2


def test_lookup_unknown_region():
    inn = make_inn12("0000000001")
    with pytest.raises(HTTPException) as exc_info:
        lookup_fl_address(inn)
    assert exc_info.value.status_code == 422
    assert "00" in exc_info.value.detail


def test_api_lookup_ok():
    res = client.get("/api/fl-address/lookup", params={"inn": "500100732259"})
    assert res.status_code == 200
    body = res.json()
    assert body["address"] == "Российская Федерация, Московская область"


def test_api_lookup_rejects_10_digits():
    res = client.get("/api/fl-address/lookup", params={"inn": "7707083893"})
    assert res.status_code == 400
    assert "юрлицо" in res.json()["detail"]


def test_inn_digits_from_excel_number():
    assert inn_digits(500100732259.0) == "500100732259"
    assert inn_digits("500100732259.0") == "500100732259"
    assert inn_digits(None) == ""
    assert inn_digits(True) == ""


def test_resolve_empty_is_skip():
    result = resolve_fl_address("")
    assert result["address"] == ""
    assert result["error"] == ""
    assert result["inn"] == ""


def test_enrich_rows_inserts_address_after_inn():
    inn = "500100732259"
    rows = [
        ["Имя", "ИНН", "Комментарий"],
        ["Иван", inn, "ок"],
        ["Пётр", "", "пусто"],
        ["ООО", "7707083893", "юл"],
    ]
    enriched, stats = enrich_rows(rows)
    assert enriched[0] == ["Имя", "ИНН", "Адрес", "Комментарий"]
    assert enriched[1][2] == "Российская Федерация, Московская область"
    assert enriched[1][3] == "ок"
    assert enriched[2][2] is None
    assert "юрлицо" in str(enriched[3][2])
    assert stats.filled == 1
    assert stats.empty == 1
    assert stats.errors == 1
    assert stats.problems[0].row == 4
    assert stats.problems[0].inn == "7707083893"


def test_enrich_rows_reuses_existing_address_column():
    inn = "500100732259"
    rows = [
        ["ИНН", "Адрес", "Код"],
        [inn, "старое", "x"],
    ]
    enriched, stats = enrich_rows(rows)
    assert enriched[0] == ["ИНН", "Адрес", "Код"]
    assert enriched[1][1] == "Российская Федерация, Московская область"
    assert stats.filled == 1


def test_enrich_rows_requires_inn_header():
    with pytest.raises(HTTPException) as exc_info:
        enrich_rows([["Имя", "Налог"], ["Иван", "500100732259"]])
    assert exc_info.value.status_code == 400
    assert "ИНН" in exc_info.value.detail


def _xlsx_bytes(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_process_batch_xlsx_keeps_other_columns():
    inn = "500100732259"
    content = _xlsx_bytes(
        [
            ["Договор", "ИНН", "Сумма"],
            ["Д-1", inn, 100],
        ]
    )
    result = process_batch_file(content, "partners.xlsx")
    assert result.filename == "fl_address_result.xlsx"
    workbook = load_workbook(io.BytesIO(result.content))
    sheet = workbook.active
    assert [cell.value for cell in sheet[1]] == ["Договор", "ИНН", "Адрес", "Сумма"]


def _batch_download(content: bytes, filename: str, output_format: str = "auto"):
    start = client.post(
        "/api/fl-address/batch",
        files={"file": (filename, content, "application/octet-stream")},
        data={"output_format": output_format},
    )
    if start.status_code != 200:
        return start, None, None
    job_id = start.json()["job_id"]
    status = client.get(f"/api/fl-address/batch/{job_id}")
    download = client.get(f"/api/fl-address/batch/{job_id}/download")
    return start, status, download


def test_api_batch_xlsx_ok():
    inn = "500100732259"
    content = _xlsx_bytes([["ИНН", "Имя"], [inn, "Иван"]])
    start, status, download = _batch_download(content, "fl.xlsx")
    assert start.status_code == 200
    assert start.json()["filled"] == 1
    assert status.status_code == 200
    assert download.status_code == 200
    assert "fl_address_result.xlsx" in download.headers["content-disposition"]
    workbook = load_workbook(io.BytesIO(download.content))
    assert workbook.active.cell(2, 2).value == "Российская Федерация, Московская область"


def test_api_batch_problem_lists_row_and_inn():
    content = _xlsx_bytes(
        [
            ["ИНН", "Имя"],
            ["500100732259", "Иван"],
            ["550509537475", "Бойко"],
        ]
    )
    start, _, download = _batch_download(content, "fl.xlsx")
    assert start.status_code == 200
    problems = start.json()["problems"]
    assert problems["total"] == 1
    assert problems["items"][0]["row"] == 3
    assert problems["items"][0]["inn"] == "550509537475"
    assert download.status_code == 200


def test_index_includes_fl_address_batch_tab():
    res = client.get("/")
    assert res.status_code == 200
    assert "fl-address-sub-batch-btn" in res.text
    assert "fl-address-panel-batch" in res.text
