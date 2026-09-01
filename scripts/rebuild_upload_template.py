"""Собрать чистый шаблон загрузки DSP (строки 1–2, лист Sheet1)."""

from pathlib import Path
import sys

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.config import TEMPLATE_PATH
from app.engine.lookups import TEMPLATE_HEADERS

DESCRIPTIONS = (
    "Токен креатива\nОбязательное поле",
    "Обязательное поле",
    "Обязательное поле",
    (
        "Обязательное поле\nДоступные значения:\n"
        "- Intermediary (Посреднический)\n- Посреднический\n"
        "- Additional (доп. соглашение)\n- Дополнительное соглашение\n"
        "- Original (оказания услуг)\n- Оказание услуг"
    ),
    (
        "Обязательное поле\nДоступные значения:\n"
        "- Representation (Представительство)\n- Представительство\n"
        "- Distribution (Договор на распространение рекламы)\n"
        "- Договор на распространение рекламы\n"
        "- DistributionOrganization (Договор на организацию распространения рекламы)\n"
        "- Договор на организацию распространения рекламы\n"
        "- Mediation (Посредничество)\n- Посредничество\n"
        "- Other (Другое)\n- Иное\n- Другое"
    ),
    (
        "Обязательное поле\nДоступные значения:\n"
        "- Conclude (Заключение договоров)\n- Заключение договоров\n"
        "- Distribution (Действия в целях распространения рекламы)\n"
        "- Действия в целях распространения рекламы\n"
        "- Commercial (Коммерческое представительство)\n"
        "- Коммерческое представительство\n"
        "- Other (другое)\n- Иное\n- Другое\n- None (нет)"
    ),
    (
        "Обязательное поле\nДоступные значения:\n"
        "-LegalPerson (юрлицо)\n-  IndividualEntrepreneur (ИП)\n"
        "- PhysicalPerson (физлицо)\n- ForeignPhysicalPerson (Иностранное физлицо)\n"
        "- ForeignLegalPerson (Иностранное юрлицо)"
    ),
    "Обязательное поле",
    "Необязательное поле",
    (
        "Обязательное поле, если тип заказчика LegalPerson, IndividualEntrepreneur, "
        "PhysicalPerson\n"
        "Обязательное поле, если тип заказчика ForeignPhysicalPerson или "
        "ForeignLegalPerson и поле Рег.номер заказчика не заполнено\n"
        "Необязательное поле, если тип заказчика ForeignPhysicalPerson или "
        "ForeignLegalPerson и поле Рег.номер заказчика заполнено"
    ),
    (
        "Обязательное поле, если тип заказчика ForeignPhysicalPerson или "
        "ForeignLegalPerson и поле ИНН заказчика не заполнено\n"
        "В противном случае необязательное поле"
    ),
    (
        "Обязательное поле, если тип заказчика ForeignPhysicalPerson или "
        "ForeignLegalPerson \nВ противном случае необязательное поле"
    ),
    (
        "Обязательное поле\nДоступные значения:\n"
        "- LegalPerson (юрлицо)\n- IndividualEntrepreneur (ИП)\n"
        "- PhysicalPerson (физлицо)\n- ForeignPhysicalPerson (Иностранное физлицо)\n"
        "- ForeignLegalPerson (Иностранное юрлицо)"
    ),
    "Обязательное поле",
    "Юридический адрес\nНеобязательное поле",
    (
        "Обязательное поле, если тип исполнителя LegalPerson, IndividualEntrepreneur, "
        "PhysicalPerson\n"
        "Обязательное поле, если тип исполнителя ForeignPhysicalPerson или "
        "ForeignLegalPerson и поле Рег.номер исполнителя не заполнено\n"
        "Необязательное поле, если тип исполнителя ForeignPhysicalPerson или "
        "ForeignLegalPerson и поле Рег.номер исполнителя заполнено"
    ),
    (
        "Обязательное поле, если тип исполнителя ForeignPhysicalPerson или "
        "ForeignLegalPerson и поле ИНН исполнителя не заполнено\n"
        "В противном случае необязательное поле"
    ),
    (
        "Обязательное поле, если тип исполнителя ForeignPhysicalPerson или "
        "ForeignLegalPerson \nВ противном случае необязательное поле"
    ),
    "Обязательное поле\nДоступные значения: \n- yes (сумма включая НДС)\n- no (сумма не включая НДС)",
    "Обязательное поле",
    "Сумма в рублях\nОбязательное поле",
)


def main() -> None:
    if len(TEMPLATE_HEADERS) != len(DESCRIPTIONS):
        raise SystemExit("TEMPLATE_HEADERS and DESCRIPTIONS length mismatch")
    path: Path = TEMPLATE_PATH
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for col, value in enumerate(TEMPLATE_HEADERS, 1):
        cell = ws.cell(1, col, value)
        cell.number_format = "@"
    for col, value in enumerate(DESCRIPTIONS, 1):
        ws.cell(2, col, value)
    ws.cell(1, 20).number_format = "0"
    ws.cell(1, 21).number_format = "#,##0.00"
    wb.save(path)
    print(f"Wrote {path} ({len(TEMPLATE_HEADERS)} columns)")


if __name__ == "__main__":
    main()
