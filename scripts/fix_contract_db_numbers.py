#!/usr/bin/env python3
"""Убирает внутренний id из колонки «Номер» в База договоров.xlsx: «26/14 - 1993» → «26/14»."""

from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import openpyxl

from app.matching.keys import norm_contract_no, strip_internal_contract_id_suffix

DEFAULT_PATH = ROOT / "Справка" / "База договоров.xlsx"
SHEET = "OriginalContract"
NUMBER_COL = 1  # A — «Номер»
TEXT_FORMAT = "@"


def fix_contract_db_numbers(path: Path = DEFAULT_PATH) -> int:
    if not path.exists():
        raise FileNotFoundError(path)

    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active
    updated = 0
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row, NUMBER_COL)
        if cell.value is None:
            continue
        if isinstance(cell.value, float) and cell.value == cell.value and cell.value == int(cell.value):
            raw = str(int(cell.value))
        else:
            raw = str(cell.value).strip()
        if not raw:
            continue
        cleaned = strip_internal_contract_id_suffix(raw)
        if cleaned != raw or cell.number_format != TEXT_FORMAT:
            cell.value = cleaned
            cell.number_format = TEXT_FORMAT
            updated += 1
    wb.save(path)
    wb.close()
    return updated


def main() -> int:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_PATH
    n = fix_contract_db_numbers(path)
    print(f"Updated {n} contract number cells in {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
