#!/usr/bin/env python3
"""Приводит колонку «Дата» в База договоров.xlsx к ISO-тексту YYYY-MM-DD для метчинга."""

from __future__ import annotations

import sys
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import openpyxl
DEFAULT_PATH = ROOT / "Справка" / "База договоров.xlsx"
SHEET = "OriginalContract"
DATE_COL = 2  # B — «Дата»
TEXT_FORMAT = "@"


def _to_iso(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, (int, float)):
        from openpyxl.utils.datetime import from_excel

        return from_excel(val).date().isoformat()
    text = str(val).strip()
    if not text:
        return None
    from app.engine.normalize import normalize_date

    return normalize_date(text) or text


def fix_contract_db_dates(path: Path = DEFAULT_PATH) -> int:
    if not path.exists():
        raise FileNotFoundError(path)

    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active
    updated = 0
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row, DATE_COL)
        iso = _to_iso(cell.value)
        if iso is None:
            continue
        if cell.value != iso or cell.number_format != TEXT_FORMAT:
            cell.value = iso
            cell.number_format = TEXT_FORMAT
            updated += 1
    wb.save(path)
    wb.close()
    return updated


def main() -> int:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_PATH
    n = fix_contract_db_dates(path)
    print(f"Updated {n} date cells in {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
