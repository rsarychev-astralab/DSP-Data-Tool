#!/usr/bin/env python3
"""Разбивает заказчика/исполнителя на имя и ИНН, переименовывает «Вид деятельности»."""

from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import openpyxl

from app.engine.template import TEXT_NUMBER_FORMAT
from app.matching.keys import split_party_name_inn

DEFAULT_PATH = ROOT / "Справка" / "База договоров.xlsx"
SHEET = "OriginalContract"

NEW_HEADERS = [
    "Номер",
    "Дата",
    "Ответственный",
    "Исполнитель",
    "ИНН исполнителя",
    "Заказчик",
    "ИНН заказчика",
    "Тип договора",
    "Сведения о предмете договора",
    "Вид деятельности",
    "Цена",
    "Включая НДС",
    "AdX",
    "Неактуален",
    "ID в OTM",
    "ID в OZON",
    "ID в VK",
]

TEXT_COLS = {2, 5, 7, 10, 13, 14, 15, 16}  # Дата, ИНН, ID


def _format_inn_cell(val: str):
    if not val:
        return None
    return str(val)


def migrate_contract_db(path: Path = DEFAULT_PATH) -> None:
    if not path.exists():
        raise FileNotFoundError(path)

    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active

    old_rows: list[list] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        old_rows.append(list(row))

    ws.delete_rows(1, ws.max_row)
    for col, title in enumerate(NEW_HEADERS, 1):
        ws.cell(1, col, title)

    for row_idx, old in enumerate(old_rows, start=2):
        contractor_name, contractor_inn = split_party_name_inn(
            old[3] if len(old) > 3 else None
        )
        customer_name, customer_inn = split_party_name_inn(
            old[4] if len(old) > 4 else None
        )
        new_row = [
            old[0],
            old[1],
            old[2] if len(old) > 2 else None,
            contractor_name or None,
            _format_inn_cell(contractor_inn),
            customer_name or None,
            _format_inn_cell(customer_inn),
            old[5] if len(old) > 5 else None,
            old[6] if len(old) > 6 else None,
            old[7] if len(old) > 7 else None,
            old[8] if len(old) > 8 else None,
            old[9] if len(old) > 9 else None,
            old[10] if len(old) > 10 else None,
            old[11] if len(old) > 11 else None,
            old[12] if len(old) > 12 else None,
            old[13] if len(old) > 13 else None,
            old[14] if len(old) > 14 else None,
        ]
        for col, val in enumerate(new_row, 1):
            if val is None:
                continue
            cell = ws.cell(row_idx, col, val)
            if col in TEXT_COLS:
                cell.number_format = TEXT_NUMBER_FORMAT

    wb.save(path)
    wb.close()


def main() -> int:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_PATH
    migrate_contract_db(path)
    print(f"Migrated {path}: {len(NEW_HEADERS)} columns, sheet {SHEET}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
